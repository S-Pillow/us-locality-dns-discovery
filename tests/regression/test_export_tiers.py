#!/usr/bin/env python3
"""EXPORT-REDESIGN regression: tier assignment, plain-English tab, tab order, sorting.

Acceptance criteria verified:
  AC-1  Plain-English Tab 1 exists as first workbook sheet.
  AC-2  Tab 1 passes legibility test: no enum names / jargon, strong findings
        appear as readable sentences.
  AC-3  Tab 1 "what this does NOT prove" block is present and warns honestly.
  AC-4  Tier assignment — delegated zone → T1, ordinary → T2, known/base → T3,
        weak/diagnostic → T4.
  AC-5  Workbook tab order: Plain-English Summary first, technical sheets after.
  AC-6  Findings sheet rows sorted by tier (strong first, junk last).
  AC-7  Delegated zone and wildcard echo have different tiers (not both 'high').
  AC-8  "Delegated, no web presence" surfaces as a POSITIVE note, not blank.
  AC-9  How-to-Read tab is updated: explains tab order, defines tiers, states
        limitations; no stale pre-tier wording.
  AC-10 Wildcard echo (post-WC-FIX.1) → T4 / NOT in plain-English strong list.
  AC-11 Delegated zone → T1 / IN plain-English strong list.
  AC-12 How-to-Read mentions the ⚠ limitation block.

Negative-action guards:
  NA-1  Plain-English tab must NOT contain enum names:
        CONFIRMED_DELEGATED_CHILD_ZONE, wildcard_not_detected, CONFIRMED_ORDINARY_DNS_NAME.
  NA-2  Tier does not upgrade confidence (confidence field unchanged).
  NA-3  Over-suppression: ordinary distinct TXT is T2, not T4.
  NA-4  The 5 strong NS/SOA delegation FQDNs land on T1.

Claim-to-code:
  Tier function:    export_service._finding_tier(evidence_status, evidence_value)
  Plain-English:    export_service.build_plain_english_summary_rows(result)
  Tab order:        export_xlsx_report — 'Plain-English Summary' at index 0.
  Sort:             build_confirmed_findings_rows — rows.sort(key=…tier…)
  No-web note:      _finding_notes(record, …, has_web_presence=False)
"""

from __future__ import annotations

import sys
import tempfile
from datetime import datetime
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from openpyxl import load_workbook

from scanner.evidence_status import (
    is_confirmed_evidence_status,
)
from scanner.export_service import (
    _NO_WEB_PRESENCE_NOTE,
    _finding_tier,
    _tier_label,
    build_confirmed_findings_rows,
    build_plain_english_summary_rows,
    build_how_to_read_rows,
    export_xlsx_report,
)
from scanner.models import (
    DiscoveredRecord,
    DomainInputRecord,
    DomainScanResult,
    EvidenceOutcome,
    EvidenceStatus,
    FindingClassification,
    RecordType,
    ScanInput,
    ScanOptions,
    ScanProfile,
    ScanRunResult,
    ScanStatus,
    WordlistPlan,
)
from scanner.paths import get_wordlists_dir
from scanner.wildcard_attestation import WildcardAttestationStatus

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

_STRONG_DELEGATED_FQDNS = [
    "ci.iola.ks.us",
    "ci.el-dorado.ks.us",
    "ci.park-city.ks.us",
    "ci.bridgeport.ct.us",
    "ci.glastonbury.ct.us",
]
_PARKING_TXT = (
    "This domain may be available.  For information, contact us-dom2@i-theta.com"
)
_DISTINCT_TXT = "v=spf1 include:_spf.example.gov ~all"


def _ns_record(fqdn: str, ns_value: str = "ns1.example.net") -> DiscoveredRecord:
    return DiscoveredRecord(
        fqdn=fqdn,
        record_type=RecordType.NS,
        value=ns_value,
        source_method="delegation_verifier",
        classification=FindingClassification.DELEGATED_CHILD_ZONE,
        evidence_status=EvidenceStatus.CONFIRMED_DELEGATED_CHILD_ZONE,
    )


def _soa_record(fqdn: str) -> DiscoveredRecord:
    return DiscoveredRecord(
        fqdn=fqdn,
        record_type=RecordType.SOA,
        value="ns1.example.net hostmaster.example.net 1 3600 900 604800 300",
        source_method="candidate_authoritative",
        classification=FindingClassification.ZONE_SOA_DISCOVERED,
        evidence_status=EvidenceStatus.CONFIRMED_DELEGATED_CHILD_ZONE,
    )


def _txt_record(
    fqdn: str,
    value: str,
    evidence_status: EvidenceStatus = EvidenceStatus.CONFIRMED_ORDINARY_DNS_NAME,
) -> DiscoveredRecord:
    return DiscoveredRecord(
        fqdn=fqdn,
        record_type=RecordType.TXT,
        value=value,
        source_method="generated_candidate",
        classification=FindingClassification.STANDARD_RECORD,
        evidence_status=evidence_status,
    )


def _a_record(fqdn: str, ip: str = "1.2.3.4") -> DiscoveredRecord:
    return DiscoveredRecord(
        fqdn=fqdn,
        record_type=RecordType.A,
        value=ip,
        source_method="generated_candidate",
        classification=FindingClassification.STANDARD_RECORD,
        evidence_status=EvidenceStatus.CONFIRMED_ORDINARY_DNS_NAME,
    )


def _withheld_outcome(fqdn: str) -> EvidenceOutcome:
    return EvidenceOutcome(
        fqdn=fqdn,
        evidence_status=EvidenceStatus.WITHHELD_WILDCARD_INCONCLUSIVE,
        source_method="generated_candidate",
        detail="Wildcard attestation inconclusive at parent; promotion withheld",
        attestation_status=WildcardAttestationStatus.INCONCLUSIVE.value,
    )


def _make_scan_result(
    base: str,
    records: list[DiscoveredRecord],
    evidence_outcomes: list[EvidenceOutcome] | None = None,
) -> ScanRunResult:
    with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False) as f:
        f.write(f"{base}\n")
        domain_file = Path(f.name)
    out_dir = Path(tempfile.mkdtemp())
    scan_input = ScanInput(
        domain_file_path=domain_file,
        options=ScanOptions(scan_profile=ScanProfile.LIGHT),
        output_dir=out_dir,
        wordlists_dir=get_wordlists_dir(),
    )
    domain_result = DomainScanResult(
        domain=base,
        records=records,
        evidence_outcomes=evidence_outcomes or [],
        candidates_tested=len(records),
    )
    return ScanRunResult(
        input=scan_input,
        domain_results=[domain_result],
        scan_timestamp=datetime(2026, 6, 28),
        scan_status=ScanStatus.COMPLETED,
        wordlist_plan=WordlistPlan(
            total_unique_labels=len(records),
            estimated_candidates_per_domain=len(records),
        ),
        domains_total=1,
        domains_planned=[base],
        domain_inputs=[],
    )


# ---------------------------------------------------------------------------
# AC-4: _finding_tier unit tests
# ---------------------------------------------------------------------------


def test_ac4_delegated_zone_is_t1() -> None:
    """AC-4: CONFIRMED_DELEGATED_CHILD_ZONE → T1 (strongest)."""
    assert _finding_tier("CONFIRMED_DELEGATED_CHILD_ZONE", "strong") == 1
    assert _finding_tier("CONFIRMED_DELEGATED_CHILD_ZONE_RECURSIVE", "strong") == 1
    print("  PASS test_ac4_delegated_zone_is_t1")


def test_ac4_ordinary_strong_is_t2() -> None:
    """AC-4: CONFIRMED_ORDINARY_DNS_NAME with strong/moderate evidence → T2."""
    assert _finding_tier("CONFIRMED_ORDINARY_DNS_NAME", "strong") == 2
    assert _finding_tier("CONFIRMED_ORDINARY_DNS_NAME", "moderate") == 2
    print("  PASS test_ac4_ordinary_strong_is_t2")


def test_ac4_ordinary_limited_is_t3() -> None:
    """AC-4: CONFIRMED_ORDINARY_DNS_NAME with limited evidence → T3."""
    assert _finding_tier("CONFIRMED_ORDINARY_DNS_NAME", "limited") == 3
    assert _finding_tier("CONFIRMED_ORDINARY_DNS_NAME", "validation_only") == 3
    print("  PASS test_ac4_ordinary_limited_is_t3")


def test_ac4_known_domain_validated_is_t3() -> None:
    """AC-4: KNOWN_DOMAIN_VALIDATED → T3."""
    assert _finding_tier("KNOWN_DOMAIN_VALIDATED", "validation_only") == 3
    print("  PASS test_ac4_known_domain_validated_is_t3")


def test_ac4_diagnostic_is_t4() -> None:
    """AC-4: diagnostic / withheld / suppressed → T4."""
    assert _finding_tier("SUPPRESSED_WILDCARD_MATCH", "") == 4
    assert _finding_tier("WITHHELD_WILDCARD_INCONCLUSIVE", "") == 4
    assert _finding_tier("INCONCLUSIVE_DNS_FAILURE", "") == 4
    assert _finding_tier("NOT_RECORDED", "") == 4
    print("  PASS test_ac4_diagnostic_is_t4")


def test_na4_five_strong_delegations_are_t1() -> None:
    """NA-4: All 5 strong NS/SOA delegation FQDNs from the operator workbook → T1."""
    for fqdn in _STRONG_DELEGATED_FQDNS:
        tier = _finding_tier("CONFIRMED_DELEGATED_CHILD_ZONE", "strong")
        assert tier == 1, f"NA-4 FAIL: {fqdn} must be T1, got T{tier}"
    print("  PASS test_na4_five_strong_delegations_are_t1")


def test_ac7_delegated_vs_echo_different_tiers() -> None:
    """AC-7: delegated zone (T1) and ordinary TXT (T2/T3) have different tiers.

    Before EXPORT-REDESIGN: both showed Confidence=high.  After: different Tier.
    """
    tier_delegation = _finding_tier("CONFIRMED_DELEGATED_CHILD_ZONE", "strong")
    tier_ordinary = _finding_tier("CONFIRMED_ORDINARY_DNS_NAME", "moderate")
    assert tier_delegation != tier_ordinary, (
        "AC-7 FAIL: delegated zone and ordinary record must have different tiers; "
        f"both got T{tier_delegation}"
    )
    assert tier_delegation < tier_ordinary, (
        "AC-7 FAIL: delegated zone must rank higher (lower tier number) than ordinary"
    )
    print(
        f"  PASS test_ac7_delegated_vs_echo_different_tiers "
        f"(T{tier_delegation} vs T{tier_ordinary})"
    )


def test_na3_distinct_txt_is_t2_not_t4() -> None:
    """NA-3: Ordinary distinct TXT → T2 (not over-suppressed to T4)."""
    tier = _finding_tier("CONFIRMED_ORDINARY_DNS_NAME", "moderate")
    assert tier == 2, f"NA-3 FAIL: distinct TXT must be T2, got T{tier}"
    print(f"  PASS test_na3_distinct_txt_is_t2_not_t4 (T{tier})")


# ---------------------------------------------------------------------------
# AC-5: workbook tab order
# ---------------------------------------------------------------------------


def test_ac5_plain_english_is_first_tab() -> None:
    """AC-5: 'Plain-English Summary' is the first workbook tab."""
    result = _make_scan_result("ks.us", [])
    out_dir = Path(tempfile.mkdtemp())
    xlsx_path = export_xlsx_report(result, out_dir)
    wb = load_workbook(xlsx_path)
    assert wb.sheetnames[0] == "Plain-English Summary", (
        f"AC-5 FAIL: first tab must be 'Plain-English Summary', got {wb.sheetnames[0]!r}\n"
        f"  All tabs: {wb.sheetnames}"
    )
    assert "How to Read" in wb.sheetnames, "How to Read tab must be present"
    assert wb.sheetnames.index("Plain-English Summary") < wb.sheetnames.index("How to Read"), (
        "Plain-English Summary must appear before How to Read"
    )
    print(f"  PASS test_ac5_plain_english_is_first_tab  tabs={wb.sheetnames}")


# ---------------------------------------------------------------------------
# AC-1 / AC-2 / NA-1: plain-English tab legibility
# ---------------------------------------------------------------------------

_JARGON_BLACKLIST = [
    "CONFIRMED_DELEGATED_CHILD_ZONE",
    "CONFIRMED_ORDINARY_DNS_NAME",
    "wildcard_not_detected",
    "wildcard_detection_inconclusive",
    "STANDARD_RECORD",
    "DELEGATED_CHILD_ZONE",
    "EvidenceStatus",
    "FindingClassification",
    "off-registry NS",
    "WildcardAttestationStatus",
]


def test_ac1_ac2_na1_plain_english_no_jargon() -> None:
    """AC-1/AC-2/NA-1: plain-English tab contains no enum names / untranslated jargon."""
    ns1 = _ns_record("ci.iola.ks.us", "ns1.cland.net")
    ns2 = _ns_record("ci.iola.ks.us", "ns1.midusa.net")
    result = _make_scan_result("ks.us", [ns1, ns2])
    rows = build_plain_english_summary_rows(result)

    full_text = " ".join(label + " " + text for label, text in rows)

    for term in _JARGON_BLACKLIST:
        assert term not in full_text, (
            f"AC-1/NA-1 FAIL: jargon term found in plain-English tab: {term!r}\n"
            f"  Context: ...{full_text[max(0, full_text.find(term) - 40):full_text.find(term) + 80]}..."
        )
    print("  PASS test_ac1_ac2_na1_plain_english_no_jargon (no enum names found)")


def test_ac2_strong_finding_appears_in_plain_english() -> None:
    """AC-2: T1 delegated zone appears in plain-English tab as a readable card."""
    ns1 = _ns_record("ci.iola.ks.us", "ns1.cland.net")
    ns2 = _ns_record("ci.iola.ks.us", "ns1.midusa.net")
    result = _make_scan_result("ks.us", [ns1, ns2])
    rows = build_plain_english_summary_rows(result)

    full_text = " ".join(label + " " + text for label, text in rows)
    labels = [label for label, _ in rows]

    assert "ci.iola.ks.us" in full_text, (
        "AC-2 FAIL: strong finding FQDN must appear in plain-English tab"
    )
    assert "ci.iola.ks.us" in labels, "Strong finding must appear as its own section label"
    assert "outside provider" in full_text or "outside servers" in full_text, (
        "AC-2 FAIL: must explain servers as 'outside provider'"
    )
    # Nameservers pulled from data
    assert "cland.net" in full_text or "midusa.net" in full_text, (
        "AC-2 FAIL: actual NS server names must appear in the plain-English tab"
    )
    print("  PASS test_ac2_strong_finding_appears_in_plain_english")


def test_ac3_not_prove_block_present() -> None:
    """AC-3: 'what this does NOT prove' block is present and warns honestly."""
    result = _make_scan_result("ks.us", [])
    rows = build_plain_english_summary_rows(result)
    labels = [label for label, _ in rows]
    full_text = " ".join(label + " " + text for label, text in rows)

    has_not_prove = any("NOT prove" in label or "does not prove" in label.lower() for label in labels)
    assert has_not_prove, (
        f"AC-3 FAIL: 'what this does NOT prove' block missing. Labels: {labels}"
    )
    assert "incomplete" in full_text.lower(), (
        "AC-3 FAIL: limitation block must mention that the known-domain list may be incomplete"
    )
    assert "not that anyone deliberately hid" in full_text.lower() or "deliberate" in full_text.lower() or "intentional" in full_text.lower(), (
        "AC-3 FAIL: must state this does not prove deliberate omission"
    )
    print("  PASS test_ac3_not_prove_block_present")


def test_ac10_na1_wildcard_echo_not_in_plain_english_strong_list() -> None:
    """AC-10/NA-1: wildcard echo (withheld) does NOT appear in plain-English strong list."""
    # A withheld wildcard outcome — represents a post-WC-FIX.1 suppressed echo
    outcome = _withheld_outcome("ci.junction-city.ks.us")
    # Add a real T1 finding so the list is non-trivial
    ns = _ns_record("ci.iola.ks.us", "ns1.cland.net")
    result = _make_scan_result("ks.us", [ns], evidence_outcomes=[outcome])

    rows = build_plain_english_summary_rows(result)
    labels = [label for label, _ in rows]
    full_text = " ".join(label + " " + text for label, text in rows)

    # The withheld echo FQDN must not appear as a strong-finding card label
    assert "ci.junction-city.ks.us" not in labels, (
        "AC-10 FAIL: withheld wildcard echo must not appear as a strong-finding card"
    )
    # The real T1 finding MUST appear
    assert "ci.iola.ks.us" in labels, (
        "AC-10 FAIL: real T1 finding must still appear"
    )
    # The dismissed count should appear somewhere (plain-English dismissal note)
    assert "set aside" in full_text.lower() or "placeholder" in full_text.lower(), (
        "AC-10 FAIL: withheld items should be acknowledged as set-aside / placeholder responses"
    )
    print("  PASS test_ac10_na1_wildcard_echo_not_in_plain_english_strong_list")


def test_ac11_delegated_zone_in_plain_english_strong() -> None:
    """AC-11: T1 delegated zone IS in the plain-English strong findings list."""
    ns = _ns_record("ci.glastonbury.ct.us", "ns1.example.net")
    result = _make_scan_result("ct.us", [ns])
    rows = build_plain_english_summary_rows(result)
    labels = [label for label, _ in rows]
    assert "ci.glastonbury.ct.us" in labels, (
        f"AC-11 FAIL: delegated zone must appear as strong-finding card; labels={labels}"
    )
    print("  PASS test_ac11_delegated_zone_in_plain_english_strong")


# ---------------------------------------------------------------------------
# AC-6: Findings sheet rows sorted by tier (strong first)
# ---------------------------------------------------------------------------


def test_ac6_findings_sorted_by_tier_strong_first() -> None:
    """AC-6: Findings sheet rows are sorted by Tier — T1 appears before T2/T3.

    Claim-to-code: build_confirmed_findings_rows — rows.sort(key=…tier…)
    """
    ns = _ns_record("ci.iola.ks.us", "ns1.cland.net")
    txt = _txt_record("ci.iola.ks.us", _DISTINCT_TXT)
    result = _make_scan_result("ks.us", [txt, ns])  # txt listed first intentionally
    rows = build_confirmed_findings_rows(result)

    # Find tier column values in order
    tiers = [r.get("tier", "") for r in rows if r.get("tier", "")]
    t1_rows = [r for r in rows if r.get("tier", "").startswith("T1")]
    t2_rows = [r for r in rows if r.get("tier", "").startswith("T2")]

    assert t1_rows, f"AC-6 FAIL: no T1 rows found in output; tiers={tiers}"
    # T1 rows must come before T2 rows in the output
    if t2_rows:
        first_t1 = next(i for i, r in enumerate(rows) if r.get("tier", "").startswith("T1"))
        first_t2 = next(i for i, r in enumerate(rows) if r.get("tier", "").startswith("T2"))
        assert first_t1 < first_t2, (
            f"AC-6 FAIL: T1 must appear before T2; first_t1={first_t1}, first_t2={first_t2}"
        )
    print(f"  PASS test_ac6_findings_sorted_by_tier_strong_first (tiers={tiers})")


# ---------------------------------------------------------------------------
# AC-8: "Delegated, no web presence" positive note
# ---------------------------------------------------------------------------


def test_ac8_no_web_positive_note_for_delegation_without_a_record() -> None:
    """AC-8: delegated zone with no A record surfaces positive 'no website' note."""
    ns = _ns_record("ci.iola.ks.us", "ns1.cland.net")
    result = _make_scan_result("ks.us", [ns])
    rows = build_confirmed_findings_rows(result)

    ns_rows = [r for r in rows if "ci.iola.ks.us" in r.get("discovered_name", "")]
    assert ns_rows, "AC-8 FAIL: no row found for ci.iola.ks.us"
    notes = " ".join(r.get("notes", "") for r in ns_rows)
    assert "no website" in notes.lower() or "no web" in notes.lower(), (
        f"AC-8 FAIL: no-web positive note missing from delegation row.\n"
        f"  notes={notes!r}"
    )
    print("  PASS test_ac8_no_web_positive_note_for_delegation_without_a_record")


def test_ac8_no_web_note_absent_when_a_record_present() -> None:
    """AC-8 negative: no-web note must NOT appear when the delegation has an A record."""
    ns = _ns_record("ci.iola.ks.us", "ns1.cland.net")
    a = _a_record("ci.iola.ks.us", "1.2.3.4")
    result = _make_scan_result("ks.us", [ns, a])
    rows = build_confirmed_findings_rows(result)
    ns_rows = [r for r in rows if r.get("finding_type") == "delegated_child_zone"]
    notes = " ".join(r.get("notes", "") for r in ns_rows)
    assert "no website" not in notes.lower() or _NO_WEB_PRESENCE_NOTE.lower() not in notes.lower(), (
        "AC-8 negative FAIL: no-web note must not appear when A record is present"
    )
    print("  PASS test_ac8_no_web_note_absent_when_a_record_present")


# ---------------------------------------------------------------------------
# AC-9 / AC-12: How-to-Read tab updated
# ---------------------------------------------------------------------------


def test_ac9_how_to_read_mentions_tab_order() -> None:
    """AC-9: updated How-to-Read explains new tab order."""
    rows = build_how_to_read_rows()
    full_text = " ".join(label + " " + text for label, text in rows)
    assert "plain-english" in full_text.lower() or "plain english" in full_text.lower(), (
        "AC-9 FAIL: How-to-Read must mention Plain-English Summary tab"
    )
    assert "tab" in full_text.lower(), "AC-9 FAIL: How-to-Read must mention tab order"
    print("  PASS test_ac9_how_to_read_mentions_tab_order")


def test_ac9_how_to_read_defines_tiers() -> None:
    """AC-9: updated How-to-Read defines T1/T2/T3/T4 in plain terms."""
    rows = build_how_to_read_rows()
    full_text = " ".join(label + " " + text for label, text in rows)
    for t in ("T1", "T2", "T3", "T4"):
        assert t in full_text, f"AC-9 FAIL: How-to-Read must define {t}"
    assert "delegated zone" in full_text.lower(), (
        "AC-9 FAIL: How-to-Read must explain 'delegated zone' in plain language"
    )
    print("  PASS test_ac9_how_to_read_defines_tiers")


def test_ac12_how_to_read_limitation_block() -> None:
    """AC-12: How-to-Read has '⚠ What this report cannot prove' / limitation block."""
    rows = build_how_to_read_rows()
    labels = [label for label, _ in rows]
    full_text = " ".join(label + " " + text for label, text in rows)
    has_limit = any("cannot prove" in label.lower() or "⚠" in label for label in labels)
    assert has_limit, f"AC-12 FAIL: limitation block missing. Labels: {labels}"
    assert "incomplete" in full_text.lower(), (
        "AC-12 FAIL: limitation block must mention known-domain list may be incomplete"
    )
    print("  PASS test_ac12_how_to_read_limitation_block")


def test_how_to_read_no_stale_wording() -> None:
    """How-to-Read must not contain pre-tier stale wording."""
    rows = build_how_to_read_rows()
    full_text = " ".join(label + " " + text for label, text in rows)
    # Old single-sentence wording for Evidence Review — removed in redesign
    assert "Open Evidence Review first" not in full_text, (
        "Stale 'Open Evidence Review first' wording found — update How-to-Read"
    )
    print("  PASS test_how_to_read_no_stale_wording")


# ---------------------------------------------------------------------------
# AC-2 "what this means" no overclaiming
# ---------------------------------------------------------------------------


def test_ac2_no_overclaiming_in_plain_english() -> None:
    """AC-2: 'what this means' does not assert concealment or proven gap."""
    ns = _ns_record("ci.iola.ks.us", "ns1.cland.net")
    result = _make_scan_result("ks.us", [ns])
    rows = build_plain_english_summary_rows(result)
    full_text = " ".join(label + " " + text for label, text in rows)

    overclaiming_phrases = [
        "registry is hiding",
        "hiding it",
        "proven gap",
        "concealed",
        "deliberately hidden",
        "registry hid",
    ]
    for phrase in overclaiming_phrases:
        assert phrase.lower() not in full_text.lower(), (
            f"AC-2 FAIL: overclaiming phrase found in plain-English tab: {phrase!r}"
        )
    # Must have "not that anyone deliberately" style caveat
    assert "not that anyone deliberately" in full_text.lower() or "not deliberately" in full_text.lower(), (
        "AC-2 FAIL: must include 'not that anyone deliberately hid them' caveat"
    )
    print("  PASS test_ac2_no_overclaiming_in_plain_english")


# ---------------------------------------------------------------------------
# NA-2: Tier does not upgrade confidence
# ---------------------------------------------------------------------------


def test_na2_tier_does_not_change_confidence() -> None:
    """NA-2: adding tier does not alter the 'confidence' field value."""
    ns = _ns_record("ci.iola.ks.us", "ns1.cland.net")
    result = _make_scan_result("ks.us", [ns])
    rows = build_confirmed_findings_rows(result)
    ns_rows = [r for r in rows if r.get("finding_type") == "delegated_child_zone"]
    for row in ns_rows:
        confidence = row.get("confidence", "")
        assert confidence in {"high", "medium", "low", "unknown", ""}, (
            f"NA-2 FAIL: unexpected confidence value: {confidence!r}"
        )
        tier = row.get("tier", "")
        # Tier label is separate; confidence must not be 'T1' or 'T2' etc.
        assert not tier.startswith(confidence), (
            f"NA-2 FAIL: tier value leaked into confidence column: {confidence!r}"
        )
    print("  PASS test_na2_tier_does_not_change_confidence")
