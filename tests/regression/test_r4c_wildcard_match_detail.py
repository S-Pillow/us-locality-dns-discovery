#!/usr/bin/env python3
"""R4c verification: Wildcard Suppression Match-Detail Capture.

Durable regression tests confirming that SUPPRESSED_WILDCARD_MATCH diagnostics
carry accurate matched_rrtype and matched_values reflecting the actual wildcard
signature entries, and that this recording is purely observational — it does not
change any promote/suppress gate decision.

Acceptance criteria verified:
  AC1 — SUPPRESSED outcome carries matched_rrtype and matched_values reflecting
         the actual matched signature entries across CSV/XLSX/JSON
  AC2 — WITHHELD outcome still has those two fields blank/not_applicable
  AC3 — All other §8 fields unchanged from R4b
  AC4 — Gate invariance: same fixture as R4a yields the same promote/suppress
         outcomes; match-detail recording is a side-channel, not a decision input

Negative-action tests:
  - matched_values contains only signature-matched values, not all candidate records
  - Inconclusive diagnostics do not gain spurious match detail
  - Non-wildcard outcomes retain blank match-detail fields

Claim-to-code:
  Engine capture: scanner/scan_engine._suppression_match_detail (after
  candidate_differentiates returns None) and the suppression branch inside
  _test_candidates where suppressed_outcome.matched_rrtype /
  suppressed_outcome.matched_values are stamped.
"""

from __future__ import annotations

import csv
import json
import sys
import tempfile
from datetime import datetime
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from tests.regression._chain import run_durable_regression
from tests.regression._paths import REGRESSION_DIR

from openpyxl import load_workbook

from scanner.evidence_status import (
    outcome_suppressed_wildcard_match,
    outcome_withheld_wildcard_inconclusive,
)
from scanner.export_service import (
    build_diagnostics_rows,
    export_csv,
    export_diagnostics_csv,
    export_json,
    export_xlsx_report,
)
from scanner.models import (
    DiscoveredRecord,
    DomainScanResult,
    EvidenceOutcome,
    EvidenceStatus,
    FindingClassification,
    RecordType,
    ScanInput,
    ScanOptions,
    ScanProfile,
    ScanRunResult,
    ScanStatus,
    WordlistPlan,
)
from scanner.paths import get_wordlists_dir
from scanner.wildcard_attestation import (
    WildcardAttestation,
    WildcardAttestationStatus,
    candidate_differentiates,
)
# Import the private helper for direct unit testing (claim-to-code §R4c).
from scanner.scan_engine import _suppression_match_detail

_BASE = "ci.lawrence.ma.us"

# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------


def _detected(type_sigs: dict[str, set[str]]) -> WildcardAttestation:
    """Build a DETECTED attestation from a dict of type -> value sets."""
    address_pool: frozenset[str] = frozenset(type_sigs.get("A", set())) | frozenset(
        type_sigs.get("AAAA", set())
    )
    return WildcardAttestation(
        status=WildcardAttestationStatus.DETECTED,
        parent=_BASE,
        type_signatures={k: frozenset(v) for k, v in type_sigs.items()},
        address_pool=address_pool,
    )


def _a_record(fqdn: str, ip: str) -> DiscoveredRecord:
    return DiscoveredRecord(
        fqdn=fqdn,
        record_type=RecordType.A,
        value=ip,
        source_method="generated_candidate",
        classification=FindingClassification.STANDARD_RECORD,
        ttl=300,
        evidence_status=EvidenceStatus.CONFIRMED_ORDINARY_DNS_NAME,
    )


def _mx_record(fqdn: str, exchange: str) -> DiscoveredRecord:
    return DiscoveredRecord(
        fqdn=fqdn,
        record_type=RecordType.MX,
        value=exchange,
        source_method="generated_candidate",
        classification=FindingClassification.STANDARD_RECORD,
        ttl=300,
    )


def _suppressed_outcome_with_detail(
    fqdn: str,
    matched_rrtype: str,
    matched_values: list[str],
) -> EvidenceOutcome:
    """Build a SUPPRESSED outcome with match detail pre-populated (as the engine would)."""
    outcome = outcome_suppressed_wildcard_match(fqdn, parent=_BASE, source_method="generated_candidate")
    outcome.matched_rrtype = matched_rrtype
    outcome.matched_values = matched_values
    return outcome


def _withheld_outcome(fqdn: str) -> EvidenceOutcome:
    return outcome_withheld_wildcard_inconclusive(fqdn, parent=_BASE, source_method="generated_candidate")


def _make_run_result(
    records: list[DiscoveredRecord],
    evidence_outcomes: list[EvidenceOutcome],
    base: str = _BASE,
) -> ScanRunResult:
    domain_result = DomainScanResult(
        domain=base,
        records=records,
        evidence_outcomes=evidence_outcomes,
        candidates_tested=len(records) + len(evidence_outcomes),
    )
    with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False) as f:
        f.write(f"{base}\n")
        domain_file = Path(f.name)
    out_dir = Path(tempfile.mkdtemp())
    scan_input = ScanInput(
        domain_file_path=domain_file,
        options=ScanOptions(scan_profile=ScanProfile.LIGHT),
        output_dir=out_dir,
        wordlists_dir=get_wordlists_dir(),
    )
    return ScanRunResult(
        input=scan_input,
        domain_results=[domain_result],
        scan_timestamp=datetime(2026, 1, 1),
        scan_status=ScanStatus.COMPLETED,
        wordlist_plan=WordlistPlan(total_unique_labels=4, estimated_candidates_per_domain=4),
        domains_total=1,
        domains_planned=[base],
        domain_inputs=[],
    )


# ===========================================================================
# 0 — Prior regression chain
# ===========================================================================


def test_prior_chain() -> None:
    """R4b regression must pass before R4c tests run."""
    run_durable_regression(REGRESSION_DIR / "test_r4b_wildcard_reporting.py")
    print("  prior chain: test_r4b_wildcard_reporting passed")


# ===========================================================================
# 1 — Unit tests on _suppression_match_detail (claim-to-code: scan_engine.py)
# ===========================================================================


def test_match_detail_a_pool_containment() -> None:
    """_suppression_match_detail: A record whose IP is in the wildcard pool → A + IP."""
    att = _detected({"A": {"1.2.3.4", "1.2.3.5"}})
    records = [_a_record(f"mail.{_BASE}", "1.2.3.4")]
    matched_type, matched_vals = _suppression_match_detail(records, att)
    assert matched_type == "A", f"Expected matched_type='A', got {matched_type!r}"
    assert "1.2.3.4" in matched_vals, f"Expected '1.2.3.4' in matched_vals; got {matched_vals}"
    print("  PASS test_match_detail_a_pool_containment (claim-to-code: _suppression_match_detail)")


def test_match_detail_mx_set_membership() -> None:
    """_suppression_match_detail: MX record whose value is in the wildcard MX set → MX + exchange."""
    att = _detected({"MX": {"mail.example.com"}})
    records = [_mx_record(f"foo.{_BASE}", "mail.example.com")]
    matched_type, matched_vals = _suppression_match_detail(records, att)
    assert matched_type == "MX", f"Expected matched_type='MX', got {matched_type!r}"
    assert "mail.example.com" in matched_vals, (
        f"Expected 'mail.example.com' in matched_vals; got {matched_vals}"
    )
    print("  PASS test_match_detail_mx_set_membership")


def test_match_detail_multiple_matching_records() -> None:
    """_suppression_match_detail: two matching A records → both IPs in matched_values."""
    att = _detected({"A": {"1.2.3.4", "1.2.3.5"}})
    records = [
        _a_record(f"mail.{_BASE}", "1.2.3.4"),
        _a_record(f"mail.{_BASE}", "1.2.3.5"),
    ]
    matched_type, matched_vals = _suppression_match_detail(records, att)
    assert matched_type == "A"
    assert "1.2.3.4" in matched_vals
    assert "1.2.3.5" in matched_vals
    print("  PASS test_match_detail_multiple_matching_records")


def test_match_detail_unmatched_record_excluded() -> None:
    """Negative-action: a record whose value is NOT in the signature is excluded from matched_values.

    This guards against _suppression_match_detail accidentally including
    unrelated candidate records.  It should only include values that actually
    appear in the wildcard signature.
    """
    att = _detected({"A": {"1.2.3.4"}})
    # "9.9.9.9" is NOT in the pool — should be excluded.
    records = [
        _a_record(f"mail.{_BASE}", "1.2.3.4"),  # matches
        _a_record(f"mail.{_BASE}", "9.9.9.9"),  # does NOT match pool
    ]
    _, matched_vals = _suppression_match_detail(records, att)
    assert "1.2.3.4" in matched_vals, "Matching IP must appear"
    assert "9.9.9.9" not in matched_vals, (
        "Non-matching IP must NOT appear in matched_values"
    )
    print("  PASS test_match_detail_unmatched_record_excluded (negative-action)")


def test_match_detail_empty_when_no_type_in_signature() -> None:
    """_suppression_match_detail returns empty when candidate type absent from signatures.

    (This is a degenerate case — the gate would not suppress here — but the
    function should not crash or return spurious values.)
    """
    att = _detected({"A": {"1.2.3.4"}})
    # Candidate is AAAA, not in signature
    aaaa = DiscoveredRecord(
        fqdn=f"mail.{_BASE}",
        record_type=RecordType.AAAA,
        value="::1",
        source_method="generated_candidate",
        classification=FindingClassification.STANDARD_RECORD,
    )
    matched_type, matched_vals = _suppression_match_detail([aaaa], att)
    assert matched_type == "", f"Expected empty matched_type, got {matched_type!r}"
    assert matched_vals == [], f"Expected empty matched_vals, got {matched_vals!r}"
    print("  PASS test_match_detail_empty_when_no_type_in_signature")


def test_match_detail_deduplication() -> None:
    """_suppression_match_detail deduplicates repeated values and types."""
    att = _detected({"A": {"1.2.3.4"}})
    records = [
        _a_record(f"mail.{_BASE}", "1.2.3.4"),
        _a_record(f"mail.{_BASE}", "1.2.3.4"),  # duplicate
    ]
    matched_type, matched_vals = _suppression_match_detail(records, att)
    assert matched_vals.count("1.2.3.4") == 1, (
        "Duplicate matched values must be de-duplicated"
    )
    assert matched_type.count("A") == 1, "Duplicate RR types must be de-duplicated"
    print("  PASS test_match_detail_deduplication")


# ===========================================================================
# 2 — EvidenceOutcome model fields
# ===========================================================================


def test_evidence_outcome_has_matched_fields() -> None:
    """EvidenceOutcome carries matched_rrtype and matched_values fields (models.py, R4c)."""
    outcome = outcome_suppressed_wildcard_match(f"mail.{_BASE}", parent=_BASE)
    assert hasattr(outcome, "matched_rrtype"), "EvidenceOutcome must have matched_rrtype"
    assert hasattr(outcome, "matched_values"), "EvidenceOutcome must have matched_values"
    # Fresh outcome: fields are None (not yet populated)
    assert outcome.matched_rrtype is None
    assert outcome.matched_values is None
    print("  PASS test_evidence_outcome_has_matched_fields")


def test_evidence_outcome_fields_assignable() -> None:
    """matched_rrtype and matched_values can be set after construction."""
    outcome = outcome_suppressed_wildcard_match(f"mail.{_BASE}", parent=_BASE)
    outcome.matched_rrtype = "A"
    outcome.matched_values = ["1.2.3.4"]
    assert outcome.matched_rrtype == "A"
    assert outcome.matched_values == ["1.2.3.4"]
    print("  PASS test_evidence_outcome_fields_assignable")


# ===========================================================================
# 3 — Gate invariance (AC4 — same suppress/promote outcomes after R4c)
# ===========================================================================


def test_gate_invariance_suppressed_still_suppressed() -> None:
    """AC4: candidate_differentiates() still returns None (suppress) for a matching IP.

    _suppression_match_detail is called AFTER the gate decision; adding it
    cannot change the outcome of candidate_differentiates().  We verify the gate
    result is unchanged by re-running candidate_differentiates with the same fixture
    that triggered suppression in R4a.

    Claim-to-code: scan_engine._test_candidates calls candidate_differentiates()
    FIRST (line ~1666) to decide promote/suppress, THEN calls
    _suppression_match_detail() (R4c addition) only on the suppress branch —
    the recording is a side-channel.
    """
    att = _detected({"A": {"1.2.3.4"}})
    records = [_a_record(f"mail.{_BASE}", "1.2.3.4")]
    # Gate decision must still be None (suppress)
    gate_result = candidate_differentiates(records, att)
    assert gate_result is None, (
        f"AC4 FAIL: gate must still return None for matching IP; got {gate_result!r}"
    )
    # Match-detail computation does NOT affect gate result
    matched_type, matched_vals = _suppression_match_detail(records, att)
    assert matched_type == "A"
    assert "1.2.3.4" in matched_vals
    # Gate result unchanged after calling _suppression_match_detail
    gate_result_after = candidate_differentiates(records, att)
    assert gate_result_after is None, (
        "AC4 FAIL: gate result must be unchanged after calling _suppression_match_detail"
    )
    print("  PASS test_gate_invariance_suppressed_still_suppressed (AC4)")


def test_gate_invariance_promoted_still_promoted() -> None:
    """AC4: distinct-IP candidate still promotes (differentiation_reason != None) after R4c."""
    att = _detected({"A": {"1.2.3.4"}})
    records = [_a_record(f"mail.{_BASE}", "5.6.7.8")]  # outside pool
    gate_result = candidate_differentiates(records, att)
    assert gate_result is not None, (
        f"AC4 FAIL: distinct-IP must still promote; got None"
    )
    print("  PASS test_gate_invariance_promoted_still_promoted (AC4)")


# ===========================================================================
# 4 — Export layer: diagnostics rows reflect match detail
# ===========================================================================


def test_diagnostics_row_matched_rrtype_populated() -> None:
    """AC1: SUPPRESSED diagnostic row has matched_rrtype set from EvidenceOutcome."""
    outcome = _suppressed_outcome_with_detail(f"mail.{_BASE}", "A", ["1.2.3.4"])
    result = _make_run_result([], [outcome])
    rows = build_diagnostics_rows(result)
    sup_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.SUPPRESSED_WILDCARD_MATCH.value]
    assert sup_rows, "Expected at least one SUPPRESSED_WILDCARD_MATCH row"
    row = sup_rows[0]
    assert row["matched_rrtype"] == "A", (
        f"AC1: matched_rrtype must be 'A', got {row['matched_rrtype']!r}"
    )
    print("  PASS test_diagnostics_row_matched_rrtype_populated (AC1)")


def test_diagnostics_row_matched_values_populated() -> None:
    """AC1: SUPPRESSED diagnostic row has matched_values (CSV: comma-separated string)."""
    outcome = _suppressed_outcome_with_detail(f"mail.{_BASE}", "A", ["1.2.3.4", "1.2.3.5"])
    result = _make_run_result([], [outcome])
    rows = build_diagnostics_rows(result)
    sup_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.SUPPRESSED_WILDCARD_MATCH.value]
    assert sup_rows
    row = sup_rows[0]
    # CSV/row format: comma-separated string
    assert "1.2.3.4" in row["matched_values"], (
        f"AC1: matched_values must contain '1.2.3.4', got {row['matched_values']!r}"
    )
    assert "1.2.3.5" in row["matched_values"], (
        f"AC1: matched_values must contain '1.2.3.5', got {row['matched_values']!r}"
    )
    print("  PASS test_diagnostics_row_matched_values_populated (AC1)")


def test_withheld_diagnostic_matched_fields_blank() -> None:
    """AC2: WITHHELD diagnostic still has blank matched_rrtype and matched_values."""
    outcome = _withheld_outcome(f"mail.{_BASE}")
    result = _make_run_result([], [outcome])
    rows = build_diagnostics_rows(result)
    wit_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.WITHHELD_WILDCARD_INCONCLUSIVE.value]
    assert wit_rows
    row = wit_rows[0]
    assert row.get("matched_rrtype", "") == "", (
        f"AC2: WITHHELD matched_rrtype must be blank, got {row.get('matched_rrtype')!r}"
    )
    assert row.get("matched_values", "") == "", (
        f"AC2: WITHHELD matched_values must be blank, got {row.get('matched_values')!r}"
    )
    print("  PASS test_withheld_diagnostic_matched_fields_blank (AC2)")


def test_suppressed_without_detail_has_empty_fields() -> None:
    """AC3: outcome with no match detail set (None) exports as empty — backward compat."""
    outcome = outcome_suppressed_wildcard_match(f"mail.{_BASE}", parent=_BASE)
    # matched_rrtype/matched_values are None (not yet populated, e.g. legacy path)
    result = _make_run_result([], [outcome])
    rows = build_diagnostics_rows(result)
    sup_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.SUPPRESSED_WILDCARD_MATCH.value]
    assert sup_rows
    row = sup_rows[0]
    assert row.get("matched_rrtype", "") == "", (
        f"Outcome with matched_rrtype=None must export as empty; got {row.get('matched_rrtype')!r}"
    )
    assert row.get("matched_values", "") == "", (
        f"Outcome with matched_values=None must export as empty; got {row.get('matched_values')!r}"
    )
    print("  PASS test_suppressed_without_detail_has_empty_fields (AC3 backward compat)")


# ===========================================================================
# 5 — CSV: diagnostics CSV includes match detail
# ===========================================================================


def test_csv_diagnostics_matched_fields() -> None:
    """AC1 CSV: diagnostics CSV rows for SUPPRESSED outcome include match detail."""
    outcome = _suppressed_outcome_with_detail(f"mail.{_BASE}", "MX", ["mail.example.com"])
    result = _make_run_result([], [outcome])
    out_dir = Path(tempfile.mkdtemp())
    diag_csv_path = export_diagnostics_csv(result, out_dir)
    with diag_csv_path.open(encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
    sup_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.SUPPRESSED_WILDCARD_MATCH.value]
    assert sup_rows, "SUPPRESSED row not found in diagnostics CSV"
    row = sup_rows[0]
    assert row["matched_rrtype"] == "MX", (
        f"CSV: matched_rrtype must be 'MX', got {row['matched_rrtype']!r}"
    )
    assert "mail.example.com" in row["matched_values"], (
        f"CSV: matched_values must contain 'mail.example.com', got {row['matched_values']!r}"
    )
    print("  PASS test_csv_diagnostics_matched_fields (AC1 CSV)")


def test_csv_withheld_matched_fields_blank() -> None:
    """AC2 CSV: WITHHELD rows in diagnostics CSV still have blank match-detail fields."""
    outcome = _withheld_outcome(f"news.{_BASE}")
    result = _make_run_result([], [outcome])
    out_dir = Path(tempfile.mkdtemp())
    diag_csv_path = export_diagnostics_csv(result, out_dir)
    with diag_csv_path.open(encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
    wit_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.WITHHELD_WILDCARD_INCONCLUSIVE.value]
    assert wit_rows
    row = wit_rows[0]
    assert row.get("matched_rrtype", "") == "", (
        f"AC2 CSV: WITHHELD matched_rrtype must be blank, got {row.get('matched_rrtype')!r}"
    )
    print("  PASS test_csv_withheld_matched_fields_blank (AC2 CSV)")


def test_findings_csv_unaffected() -> None:
    """AC3: confirmed findings CSV does NOT gain matched_rrtype/matched_values columns."""
    from scanner.models import EvidenceStatus as ES
    rec = DiscoveredRecord(
        fqdn=f"www.{_BASE}",
        record_type=RecordType.A,
        value="10.0.0.1",
        source_method="generated_candidate",
        classification=FindingClassification.STANDARD_RECORD,
        ttl=300,
        evidence_status=ES.CONFIRMED_ORDINARY_DNS_NAME,
    )
    result = _make_run_result([rec], [])
    out_dir = Path(tempfile.mkdtemp())
    csv_path, _ = export_csv(result, out_dir)
    with csv_path.open(encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        list(reader)
    # These §8-only columns should NOT appear in the confirmed findings CSV.
    assert reader.fieldnames is not None
    assert "matched_rrtype" not in reader.fieldnames, (
        "matched_rrtype must NOT appear in confirmed findings CSV"
    )
    assert "matched_values" not in reader.fieldnames, (
        "matched_values must NOT appear in confirmed findings CSV"
    )
    print("  PASS test_findings_csv_unaffected (AC3)")


# ===========================================================================
# 6 — XLSX: Diagnostics sheet
# ===========================================================================


def test_xlsx_diagnostics_matched_values() -> None:
    """AC1 XLSX: Diagnostics sheet has correct matched_rrtype and Matched values for SUPPRESSED."""
    outcome = _suppressed_outcome_with_detail(f"mail.{_BASE}", "A", ["1.2.3.4"])
    result = _make_run_result([], [outcome])
    out_dir = Path(tempfile.mkdtemp())
    xlsx_path = export_xlsx_report(result, out_dir)
    wb = load_workbook(xlsx_path, read_only=True)
    assert "Diagnostics" in wb.sheetnames
    ws = wb["Diagnostics"]
    rows_raw = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows_raw[0]]
    data_rows = [
        dict(zip(headers, [str(v) if v is not None else "" for v in row]))
        for row in rows_raw[1:]
    ]
    sup_rows = [r for r in data_rows if "SUPPRESSED_WILDCARD_MATCH" in r.get("Evidence status", "")]
    assert sup_rows, "SUPPRESSED row not found in XLSX Diagnostics sheet"
    row = sup_rows[0]
    assert row.get("Matched RR type") == "A", (
        f"XLSX: Matched RR type must be 'A', got {row.get('Matched RR type')!r}"
    )
    assert "1.2.3.4" in row.get("Matched values", ""), (
        f"XLSX: Matched values must contain '1.2.3.4', got {row.get('Matched values')!r}"
    )
    print("  PASS test_xlsx_diagnostics_matched_values (AC1 XLSX)")


def test_xlsx_withheld_matched_fields_blank() -> None:
    """AC2 XLSX: WITHHELD row in Diagnostics sheet has blank matched detail."""
    outcome = _withheld_outcome(f"news.{_BASE}")
    result = _make_run_result([], [outcome])
    out_dir = Path(tempfile.mkdtemp())
    xlsx_path = export_xlsx_report(result, out_dir)
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Diagnostics"]
    rows_raw = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows_raw[0]]
    data_rows = [
        dict(zip(headers, [str(v) if v is not None else "" for v in row]))
        for row in rows_raw[1:]
    ]
    wit_rows = [r for r in data_rows if "WITHHELD" in r.get("Evidence status", "")]
    assert wit_rows
    row = wit_rows[0]
    assert row.get("Matched RR type", "") == "", (
        f"XLSX: WITHHELD Matched RR type must be blank, got {row.get('Matched RR type')!r}"
    )
    print("  PASS test_xlsx_withheld_matched_fields_blank (AC2 XLSX)")


# ===========================================================================
# 7 — JSON: matched_values as array; matched_rrtype as string
# ===========================================================================


def test_json_matched_values_is_list() -> None:
    """AC1 JSON: SUPPRESSED evidence_diagnostic carries matched_values as a JSON array."""
    outcome = _suppressed_outcome_with_detail(f"mail.{_BASE}", "A", ["1.2.3.4", "1.2.3.5"])
    result = _make_run_result([], [outcome])
    out_dir = Path(tempfile.mkdtemp())
    json_path = export_json(result, out_dir)
    doc = json.loads(json_path.read_text(encoding="utf-8"))
    diags = doc["domains"][0].get("evidence_diagnostics", [])
    sup_diags = [d for d in diags if d.get("tested_name") == f"mail.{_BASE}"]
    assert sup_diags, f"mail.{_BASE} not found in JSON evidence_diagnostics"
    diag = sup_diags[0]
    assert diag.get("matched_rrtype") == "A", (
        f"JSON: matched_rrtype must be 'A', got {diag.get('matched_rrtype')!r}"
    )
    matched_vals = diag.get("matched_values")
    assert isinstance(matched_vals, list), (
        f"JSON: matched_values must be a list; got {type(matched_vals).__name__}: {matched_vals!r}"
    )
    assert "1.2.3.4" in matched_vals
    assert "1.2.3.5" in matched_vals
    print("  PASS test_json_matched_values_is_list (AC1 JSON)")


def test_json_withheld_matched_fields_null() -> None:
    """AC2 JSON: WITHHELD evidence_diagnostic has matched_rrtype=None, matched_values=None."""
    outcome = _withheld_outcome(f"news.{_BASE}")
    result = _make_run_result([], [outcome])
    out_dir = Path(tempfile.mkdtemp())
    json_path = export_json(result, out_dir)
    doc = json.loads(json_path.read_text(encoding="utf-8"))
    diags = doc["domains"][0].get("evidence_diagnostics", [])
    wit_diags = [d for d in diags if d.get("tested_name") == f"news.{_BASE}"]
    assert wit_diags
    diag = wit_diags[0]
    assert diag.get("matched_rrtype") is None, (
        f"AC2 JSON: WITHHELD matched_rrtype must be None, got {diag.get('matched_rrtype')!r}"
    )
    assert diag.get("matched_values") is None, (
        f"AC2 JSON: WITHHELD matched_values must be None, got {diag.get('matched_values')!r}"
    )
    print("  PASS test_json_withheld_matched_fields_null (AC2 JSON)")


def test_json_matched_values_only_signature_entries() -> None:
    """Negative-action JSON: matched_values contains only signature-matched values.

    We set matched_values explicitly to ['1.2.3.4'] on the outcome (what the engine
    would produce) and verify the export does NOT include '9.9.9.9' (an unrelated
    candidate record value that was not in the wildcard signature).
    """
    outcome = _suppressed_outcome_with_detail(f"mail.{_BASE}", "A", ["1.2.3.4"])
    result = _make_run_result([], [outcome])
    out_dir = Path(tempfile.mkdtemp())
    json_path = export_json(result, out_dir)
    doc = json.loads(json_path.read_text(encoding="utf-8"))
    diags = doc["domains"][0].get("evidence_diagnostics", [])
    diag = [d for d in diags if d.get("tested_name") == f"mail.{_BASE}"][0]
    matched_vals = diag.get("matched_values", [])
    assert "9.9.9.9" not in matched_vals, (
        "matched_values must not contain unrelated values"
    )
    assert "1.2.3.4" in matched_vals, (
        "matched_values must contain the actual matched IP"
    )
    print("  PASS test_json_matched_values_only_signature_entries (negative-action JSON)")


# ===========================================================================
# 8 — §8 fields unchanged (AC3)
# ===========================================================================


def test_s8_other_fields_unchanged() -> None:
    """AC3: all other §8 fields on SUPPRESSED diagnostic are unchanged from R4b."""
    outcome = _suppressed_outcome_with_detail(f"mail.{_BASE}", "A", ["1.2.3.4"])
    result = _make_run_result([], [outcome])
    rows = build_diagnostics_rows(result)
    sup_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.SUPPRESSED_WILDCARD_MATCH.value]
    assert sup_rows
    row = sup_rows[0]
    assert row["wildcard_attestation_status"] == "wildcard_detected"
    assert row["wildcard_signature_matched"] == "true"
    assert row["diagnostic_reason"] == "suppressed_as_wildcard_only"
    assert row["wildcard_parent_scope"] == _BASE
    print("  PASS test_s8_other_fields_unchanged (AC3)")


# ===========================================================================
# Main
# ===========================================================================


def main() -> None:
    print("=== R4c Wildcard Suppression Match-Detail Capture Regression ===")
    test_prior_chain()

    print("\n--- _suppression_match_detail unit tests (claim-to-code) ---")
    test_match_detail_a_pool_containment()
    test_match_detail_mx_set_membership()
    test_match_detail_multiple_matching_records()
    test_match_detail_unmatched_record_excluded()
    test_match_detail_empty_when_no_type_in_signature()
    test_match_detail_deduplication()

    print("\n--- EvidenceOutcome model fields ---")
    test_evidence_outcome_has_matched_fields()
    test_evidence_outcome_fields_assignable()

    print("\n--- Gate invariance (AC4) ---")
    test_gate_invariance_suppressed_still_suppressed()
    test_gate_invariance_promoted_still_promoted()

    print("\n--- Export: diagnostics rows ---")
    test_diagnostics_row_matched_rrtype_populated()
    test_diagnostics_row_matched_values_populated()
    test_withheld_diagnostic_matched_fields_blank()
    test_suppressed_without_detail_has_empty_fields()

    print("\n--- CSV ---")
    test_csv_diagnostics_matched_fields()
    test_csv_withheld_matched_fields_blank()
    test_findings_csv_unaffected()

    print("\n--- XLSX ---")
    test_xlsx_diagnostics_matched_values()
    test_xlsx_withheld_matched_fields_blank()

    print("\n--- JSON ---")
    test_json_matched_values_is_list()
    test_json_withheld_matched_fields_null()
    test_json_matched_values_only_signature_entries()

    print("\n--- AC3: other §8 fields unchanged ---")
    test_s8_other_fields_unchanged()

    print("\n=== R4c: all assertions passed ===")


if __name__ == "__main__":
    main()
