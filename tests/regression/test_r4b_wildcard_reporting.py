#!/usr/bin/env python3
"""R4b verification: Wildcard Reporting Contract.

Durable regression tests confirming all §7–10.8 acceptance criteria, their
negative-action guards, and §10 sheet routing.

Contract §7 — six wildcard attestation fields on every confirmed finding:
  wildcard_attestation_status, wildcard_parent_scope, wildcard_probe_count,
  wildcard_rrtypes_tested, wildcard_signature_matched, wildcard_differentiation_reason

Contract §7 valid combinations:
  - wildcard_not_detected  -> signature_matched=false, reason=not_applicable
  - wildcard_detected (promoted) -> signature_matched=false, reason in allowed set
  - wildcard_inconclusive MUST NOT appear on confirmed findings

Contract §8 — wildcard diagnostic fields on SUPPRESSED/WITHHELD outcomes:
  candidate_name, parent_scope, candidate_source,
  wildcard_attestation_status, wildcard_signature_matched,
  matched_rrtype, matched_values, diagnostic_reason

Contract §9 — report wording for three cases:
  Passed, Suppressed, Inconclusive (exact phrasings from constants)

Contract §10.7 — wildcard diagnostics on Diagnostics sheet only, never Findings.
Contract §10.8 — every confirmed finding row includes the §7 fields.

Negative-action tests (required, durable):
  - No confirmed finding carries wildcard_inconclusive
  - Suppressed candidate appears ONLY in diagnostics with diagnostic_reason=suppressed_as_wildcard_only
  - A not_detected finding emits reason=not_applicable (not a differentiation reason)
  - JSON, CSV, and XLSX all checked.
"""

from __future__ import annotations

import csv
import json
import sys
import tempfile
from datetime import datetime
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from tests.regression._chain import run_durable_regression
from tests.regression._paths import REGRESSION_DIR

from openpyxl import load_workbook

from scanner.evidence_status import (
    outcome_suppressed_wildcard_match,
    outcome_withheld_wildcard_inconclusive,
)
from scanner.export_service import (
    WILDCARD_WORDING_DETECTED_PROMOTED,
    WILDCARD_WORDING_INCONCLUSIVE,
    WILDCARD_WORDING_SUPPRESSED,
    build_confirmed_findings_rows,
    build_diagnostics_rows,
    build_json_document,
    export_csv,
    export_diagnostics_csv,
    export_json,
    export_xlsx_report,
)
from scanner.models import (
    DiscoveredRecord,
    DomainInputRecord,
    DomainScanResult,
    EvidenceOutcome,
    EvidenceStatus,
    FindingClassification,
    RecordType,
    ScanInput,
    ScanOptions,
    ScanProfile,
    ScanRunResult,
    ScanStatus,
    WordlistPlan,
)
from scanner.paths import get_wordlists_dir
from scanner.wildcard_attestation import (
    ATTESTATION_PROBE_TYPES,
    MIN_PROBE_COUNT,
    REASON_DISTINCT_ANSWER,
    REASON_DISTINCT_RRTYPE,
    REASON_NO_WILDCARD,
)

# ---------------------------------------------------------------------------
# Test fixtures
# ---------------------------------------------------------------------------

_BASE = "ci.lawrence.ma.us"


def _confirmed_clean(fqdn: str) -> DiscoveredRecord:
    """A confirmed A-record finding with NO wildcard detected."""
    return DiscoveredRecord(
        fqdn=fqdn,
        record_type=RecordType.A,
        value="10.0.0.1",
        source_method="generated_candidate",
        classification=FindingClassification.STANDARD_RECORD,
        ttl=300,
        evidence_status=EvidenceStatus.CONFIRMED_ORDINARY_DNS_NAME,
        # attestation_status=None → treated as CLEAN by export layer
    )


def _confirmed_detected_promoted(fqdn: str, reason: str = REASON_DISTINCT_ANSWER) -> DiscoveredRecord:
    """A confirmed finding that differentiated despite a wildcard being DETECTED."""
    return DiscoveredRecord(
        fqdn=fqdn,
        record_type=RecordType.A,
        value="10.1.1.1",
        source_method="generated_candidate",
        classification=FindingClassification.STANDARD_RECORD,
        ttl=300,
        evidence_status=EvidenceStatus.CONFIRMED_ORDINARY_DNS_NAME,
        attestation_status="detected",
        wildcard_signature_matched=False,
        wildcard_differentiation_reason=reason,
    )


def _suppressed(fqdn: str) -> EvidenceOutcome:
    return outcome_suppressed_wildcard_match(fqdn, parent=_BASE, source_method="generated_candidate")


def _withheld(fqdn: str) -> EvidenceOutcome:
    return outcome_withheld_wildcard_inconclusive(fqdn, parent=_BASE, source_method="generated_candidate")


def _make_run_result(
    records: list[DiscoveredRecord],
    evidence_outcomes: list[EvidenceOutcome],
    base: str = _BASE,
) -> ScanRunResult:
    domain_result = DomainScanResult(
        domain=base,
        records=records,
        evidence_outcomes=evidence_outcomes,
        candidates_tested=len(records) + len(evidence_outcomes),
    )
    with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False) as f:
        f.write(f"{base}\n")
        domain_file = Path(f.name)
    out_dir = Path(tempfile.mkdtemp())
    scan_input = ScanInput(
        domain_file_path=domain_file,
        options=ScanOptions(scan_profile=ScanProfile.LIGHT),
        output_dir=out_dir,
        wordlists_dir=get_wordlists_dir(),
    )
    return ScanRunResult(
        input=scan_input,
        domain_results=[domain_result],
        scan_timestamp=datetime(2026, 1, 1),
        scan_status=ScanStatus.COMPLETED,
        wordlist_plan=WordlistPlan(total_unique_labels=4, estimated_candidates_per_domain=4),
        domains_total=1,
        domains_planned=[base],
        domain_inputs=[],
    )


# ===========================================================================
# 0 — Prior regression chain
# ===========================================================================


def test_prior_chain() -> None:
    """R4a regression must pass before R4b tests run."""
    run_durable_regression(REGRESSION_DIR / "test_r4a_wildcard_attestation.py")
    print("  prior chain: test_r4a_wildcard_attestation passed")


# ===========================================================================
# 1 — §7 finding fields on confirmed findings
# ===========================================================================


def test_s7_fields_present_on_clean_finding() -> None:
    """§10.8: every confirmed finding row includes all six §7 wildcard fields."""
    rec = _confirmed_clean(f"www.{_BASE}")
    result = _make_run_result([rec], [])
    rows = build_confirmed_findings_rows(result)
    assert rows, "Expected at least one confirmed finding row"
    row = rows[0]
    required = {
        "wildcard_attestation_status",
        "wildcard_parent_scope",
        "wildcard_probe_count",
        "wildcard_rrtypes_tested",
        "wildcard_signature_matched",
        "wildcard_differentiation_reason",
    }
    missing = required - row.keys()
    assert not missing, f"§10.8: missing §7 keys: {missing}"
    print("  PASS test_s7_fields_present_on_clean_finding (§10.8)")


def test_s7_valid_combination_not_detected() -> None:
    """§7: wildcard_not_detected -> signature_matched=false, reason=not_applicable."""
    rec = _confirmed_clean(f"www.{_BASE}")
    result = _make_run_result([rec], [])
    rows = build_confirmed_findings_rows(result)
    row = rows[0]
    assert row["wildcard_attestation_status"] == "wildcard_not_detected", (
        f"Expected wildcard_not_detected, got {row['wildcard_attestation_status']!r}"
    )
    assert row["wildcard_signature_matched"] == "false", (
        f"§7: not_detected -> signature_matched must be false, got {row['wildcard_signature_matched']!r}"
    )
    assert row["wildcard_differentiation_reason"] == "not_applicable", (
        f"§7: not_detected -> reason must be not_applicable, got {row['wildcard_differentiation_reason']!r}"
    )
    print("  PASS test_s7_valid_combination_not_detected")


def test_s7_valid_combination_detected_promoted() -> None:
    """§7: wildcard_detected (promoted) -> signature_matched=false, reason in allowed set."""
    allowed_reasons = {
        "distinct_rrtype",
        "distinct_answer",
        "distinct_cname_target",
        "candidate_ns_soa",
        "verified_delegation",
    }
    rec = _confirmed_detected_promoted(f"mail.{_BASE}", reason=REASON_DISTINCT_ANSWER)
    result = _make_run_result([rec], [])
    rows = build_confirmed_findings_rows(result)
    row = rows[0]
    assert row["wildcard_attestation_status"] == "wildcard_detected", (
        f"Expected wildcard_detected, got {row['wildcard_attestation_status']!r}"
    )
    assert row["wildcard_signature_matched"] == "false", (
        f"§7: detected+promoted -> signature_matched must be false, got {row['wildcard_signature_matched']!r}"
    )
    assert row["wildcard_differentiation_reason"] in allowed_reasons, (
        f"§7: reason {row['wildcard_differentiation_reason']!r} not in allowed set {allowed_reasons}"
    )
    print("  PASS test_s7_valid_combination_detected_promoted")


def test_s7_probe_count_and_rrtypes() -> None:
    """§7: wildcard_probe_count and wildcard_rrtypes_tested match engine constants."""
    rec = _confirmed_clean(f"www.{_BASE}")
    result = _make_run_result([rec], [])
    rows = build_confirmed_findings_rows(result)
    row = rows[0]
    assert row["wildcard_probe_count"] == str(MIN_PROBE_COUNT), (
        f"Expected probe_count={MIN_PROBE_COUNT}, got {row['wildcard_probe_count']!r}"
    )
    expected_rrtypes = ",".join(ATTESTATION_PROBE_TYPES)
    assert row["wildcard_rrtypes_tested"] == expected_rrtypes, (
        f"Expected rrtypes_tested={expected_rrtypes!r}, got {row['wildcard_rrtypes_tested']!r}"
    )
    print("  PASS test_s7_probe_count_and_rrtypes")


def test_s7_parent_scope_derived_from_fqdn() -> None:
    """§7: wildcard_parent_scope is the first label stripped from tested_name."""
    fqdn = f"www.{_BASE}"
    rec = _confirmed_clean(fqdn)
    result = _make_run_result([rec], [])
    rows = build_confirmed_findings_rows(result)
    row = rows[0]
    expected_parent = _BASE  # strip "www."
    assert row["wildcard_parent_scope"] == expected_parent, (
        f"Expected parent_scope={expected_parent!r}, got {row['wildcard_parent_scope']!r}"
    )
    print("  PASS test_s7_parent_scope_derived_from_fqdn")


# ===========================================================================
# 2 — §7 valid-combinations negative-action: no confirmed finding carries wildcard_inconclusive
# ===========================================================================


def test_no_confirmed_finding_carries_wildcard_inconclusive() -> None:
    """Negative-action §7: no confirmed finding row may carry wildcard_inconclusive.

    The promotion gate withholds INCONCLUSIVE-attestation candidates (R4a);
    build_confirmed_findings_rows must never emit a row with wildcard_inconclusive.
    Both clean and DETECTED-promoted confirmed findings are checked.
    """
    records = [
        _confirmed_clean(f"www.{_BASE}"),
        _confirmed_detected_promoted(f"mail.{_BASE}"),
    ]
    result = _make_run_result(records, [])
    rows = build_confirmed_findings_rows(result)
    for row in rows:
        att = row.get("wildcard_attestation_status", "")
        assert att != "wildcard_inconclusive", (
            f"Negative-action §7: confirmed finding for {row.get('tested_name')!r} "
            f"carries wildcard_inconclusive — must never happen"
        )
    print("  PASS test_no_confirmed_finding_carries_wildcard_inconclusive (negative-action §7)")


# ===========================================================================
# 3 — §8 diagnostic fields on wildcard diagnostics
# ===========================================================================


def test_s8_fields_on_suppressed_outcome() -> None:
    """§8: SUPPRESSED outcome carries wildcard_attestation_status=wildcard_detected,
    signature_matched=true, diagnostic_reason=suppressed_as_wildcard_only."""
    sup = _suppressed(f"mail.{_BASE}")
    result = _make_run_result([], [sup])
    rows = build_diagnostics_rows(result)
    wc_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.SUPPRESSED_WILDCARD_MATCH.value]
    assert wc_rows, "Expected at least one SUPPRESSED_WILDCARD_MATCH diagnostic row"
    row = wc_rows[0]
    assert row["wildcard_attestation_status"] == "wildcard_detected", (
        f"§8: suppressed -> wildcard_attestation_status must be wildcard_detected, got {row['wildcard_attestation_status']!r}"
    )
    assert row["wildcard_signature_matched"] == "true", (
        f"§8: suppressed -> signature_matched must be true, got {row['wildcard_signature_matched']!r}"
    )
    assert row["diagnostic_reason"] == "suppressed_as_wildcard_only", (
        f"§8: suppressed -> diagnostic_reason must be suppressed_as_wildcard_only, got {row['diagnostic_reason']!r}"
    )
    print("  PASS test_s8_fields_on_suppressed_outcome (§8)")


def test_s8_fields_on_withheld_outcome() -> None:
    """§8: WITHHELD outcome carries wildcard_attestation_status=wildcard_inconclusive,
    signature_matched=not_applicable, diagnostic_reason=wildcard_attestation_inconclusive."""
    wit = _withheld(f"mail.{_BASE}")
    result = _make_run_result([], [wit])
    rows = build_diagnostics_rows(result)
    wc_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.WITHHELD_WILDCARD_INCONCLUSIVE.value]
    assert wc_rows, "Expected at least one WITHHELD_WILDCARD_INCONCLUSIVE diagnostic row"
    row = wc_rows[0]
    assert row["wildcard_attestation_status"] == "wildcard_inconclusive", (
        f"§8: withheld -> wildcard_attestation_status must be wildcard_inconclusive, got {row['wildcard_attestation_status']!r}"
    )
    assert row["wildcard_signature_matched"] == "not_applicable", (
        f"§8: withheld -> signature_matched must be not_applicable, got {row['wildcard_signature_matched']!r}"
    )
    assert row["diagnostic_reason"] == "wildcard_attestation_inconclusive", (
        f"§8: withheld -> diagnostic_reason must be wildcard_attestation_inconclusive, got {row['diagnostic_reason']!r}"
    )
    print("  PASS test_s8_fields_on_withheld_outcome (§8)")


def test_s8_parent_scope_on_diagnostic_row() -> None:
    """§8: parent_scope field is set to parent of the diagnostic candidate."""
    fqdn = f"mail.{_BASE}"
    sup = _suppressed(fqdn)
    result = _make_run_result([], [sup])
    rows = build_diagnostics_rows(result)
    wc_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.SUPPRESSED_WILDCARD_MATCH.value]
    assert wc_rows
    row = wc_rows[0]
    assert row["wildcard_parent_scope"] == _BASE, (
        f"§8: wildcard_parent_scope expected {_BASE!r}, got {row['wildcard_parent_scope']!r}"
    )
    print("  PASS test_s8_parent_scope_on_diagnostic_row (§8)")


def test_s8_non_wildcard_outcome_has_blank_wildcard_fields() -> None:
    """§8: non-wildcard outcome rows have empty wildcard fields (no contamination)."""
    from scanner.evidence_status import outcome_skipped_by_parent_gating
    skipped = EvidenceOutcome(
        fqdn=f"skip.{_BASE}",
        evidence_status=EvidenceStatus.SKIPPED_BY_PARENT_GATING,
        source_method="generated_candidate",
        detail="parent gated",
    )
    result = _make_run_result([], [skipped])
    rows = build_diagnostics_rows(result)
    skip_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.SKIPPED_BY_PARENT_GATING.value]
    assert skip_rows, "Expected at least one SKIPPED_BY_PARENT_GATING row"
    row = skip_rows[0]
    assert row.get("wildcard_attestation_status", "") == "", (
        f"Non-wildcard row must have empty wildcard_attestation_status, got {row.get('wildcard_attestation_status')!r}"
    )
    assert row.get("diagnostic_reason", "") == "", (
        f"Non-wildcard row must have empty diagnostic_reason, got {row.get('diagnostic_reason')!r}"
    )
    print("  PASS test_s8_non_wildcard_outcome_has_blank_wildcard_fields")


# ===========================================================================
# 4 — §9 report wording
# ===========================================================================


def test_s9_wording_detected_promoted_in_notes() -> None:
    """§9: confirmed DETECTED+promoted finding has the 'Passed' contract phrasing in notes."""
    rec = _confirmed_detected_promoted(f"mail.{_BASE}")
    result = _make_run_result([rec], [])
    rows = build_confirmed_findings_rows(result)
    row = rows[0]
    assert WILDCARD_WORDING_DETECTED_PROMOTED in row["notes"], (
        f"§9: expected '{WILDCARD_WORDING_DETECTED_PROMOTED}' in notes, got: {row['notes']!r}"
    )
    print("  PASS test_s9_wording_detected_promoted_in_notes (§9)")


def test_s9_clean_finding_has_no_wildcard_detected_wording() -> None:
    """§9 negative: CLEAN finding must NOT carry the DETECTED-promoted wording."""
    rec = _confirmed_clean(f"www.{_BASE}")
    result = _make_run_result([rec], [])
    rows = build_confirmed_findings_rows(result)
    row = rows[0]
    assert WILDCARD_WORDING_DETECTED_PROMOTED not in row["notes"], (
        f"§9: CLEAN finding must not carry detected-promoted wording; notes={row['notes']!r}"
    )
    print("  PASS test_s9_clean_finding_has_no_wildcard_detected_wording (§9 negative)")


def test_s9_wording_suppressed_in_diagnostic_notes() -> None:
    """§9: SUPPRESSED diagnostic has the 'Suppressed' contract phrasing in notes."""
    sup = _suppressed(f"mail.{_BASE}")
    result = _make_run_result([], [sup])
    rows = build_diagnostics_rows(result)
    wc_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.SUPPRESSED_WILDCARD_MATCH.value]
    assert wc_rows
    row = wc_rows[0]
    assert WILDCARD_WORDING_SUPPRESSED in row["notes"], (
        f"§9: expected suppressed wording in notes, got: {row['notes']!r}"
    )
    print("  PASS test_s9_wording_suppressed_in_diagnostic_notes (§9)")


def test_s9_wording_inconclusive_in_diagnostic_notes() -> None:
    """§9: WITHHELD diagnostic has the 'Inconclusive' contract phrasing in notes."""
    wit = _withheld(f"mail.{_BASE}")
    result = _make_run_result([], [wit])
    rows = build_diagnostics_rows(result)
    wc_rows = [r for r in rows if r.get("evidence_status") == EvidenceStatus.WITHHELD_WILDCARD_INCONCLUSIVE.value]
    assert wc_rows
    row = wc_rows[0]
    assert WILDCARD_WORDING_INCONCLUSIVE in row["notes"], (
        f"§9: expected inconclusive wording in notes, got: {row['notes']!r}"
    )
    print("  PASS test_s9_wording_inconclusive_in_diagnostic_notes (§9)")


# ===========================================================================
# 5 — §10.7 sheet routing: wildcard diagnostics on Diagnostics only
# ===========================================================================


def test_s10_7_suppressed_only_in_diagnostics_not_findings() -> None:
    """§10.7 negative-action: suppressed candidate appears ONLY in diagnostics, never findings."""
    sup = _suppressed(f"mail.{_BASE}")
    result = _make_run_result([], [sup])
    confirmed_rows = build_confirmed_findings_rows(result)
    diag_rows = build_diagnostics_rows(result)

    confirmed_names = [r.get("tested_name") for r in confirmed_rows]
    diag_names = [r.get("tested_name") for r in diag_rows]

    assert f"mail.{_BASE}" not in confirmed_names, (
        f"§10.7: suppressed candidate must NOT appear on Findings sheet"
    )
    assert f"mail.{_BASE}" in diag_names, (
        f"§10.7: suppressed candidate must appear on Diagnostics sheet"
    )
    # Verify diagnostic_reason on the row
    sup_rows = [r for r in diag_rows if r.get("tested_name") == f"mail.{_BASE}"]
    assert sup_rows[0].get("diagnostic_reason") == "suppressed_as_wildcard_only", (
        f"§10.7: diagnostic_reason must be suppressed_as_wildcard_only"
    )
    print("  PASS test_s10_7_suppressed_only_in_diagnostics_not_findings (§10.7 negative-action)")


def test_s10_7_withheld_only_in_diagnostics_not_findings() -> None:
    """§10.7: withheld-inconclusive candidate appears ONLY in diagnostics."""
    wit = _withheld(f"mail.{_BASE}")
    result = _make_run_result([], [wit])
    confirmed_rows = build_confirmed_findings_rows(result)
    diag_rows = build_diagnostics_rows(result)

    confirmed_names = [r.get("tested_name") for r in confirmed_rows]
    assert f"mail.{_BASE}" not in confirmed_names, (
        f"§10.7: withheld candidate must NOT appear on Findings sheet"
    )
    wit_rows = [r for r in diag_rows if r.get("tested_name") == f"mail.{_BASE}"]
    assert wit_rows, "§10.7: withheld candidate must appear in diagnostics"
    assert wit_rows[0].get("diagnostic_reason") == "wildcard_attestation_inconclusive", (
        f"§10.7: diagnostic_reason must be wildcard_attestation_inconclusive"
    )
    print("  PASS test_s10_7_withheld_only_in_diagnostics_not_findings (§10.7)")


# ===========================================================================
# 6 — §10.8 CSV file: all §7 columns present and correctly valued
# ===========================================================================


def test_csv_s7_fields_correct() -> None:
    """§10.8 CSV: confirmed findings CSV includes all §7 fields with correct values."""
    records = [
        _confirmed_clean(f"www.{_BASE}"),
        _confirmed_detected_promoted(f"mail.{_BASE}", reason=REASON_DISTINCT_RRTYPE),
    ]
    result = _make_run_result(records, [])
    out_dir = Path(tempfile.mkdtemp())
    csv_path, _ = export_csv(result, out_dir)
    with csv_path.open(encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
    assert rows, "CSV must contain rows"
    # Check all §7 column headers present
    expected_cols = {
        "wildcard_attestation_status",
        "wildcard_parent_scope",
        "wildcard_probe_count",
        "wildcard_rrtypes_tested",
        "wildcard_signature_matched",
        "wildcard_differentiation_reason",
    }
    assert reader.fieldnames is not None
    present = set(reader.fieldnames)
    missing = expected_cols - present
    assert not missing, f"CSV missing §7 columns: {missing}"
    # Validate the CLEAN row
    clean_rows = [r for r in rows if r.get("tested_name") == f"www.{_BASE}"]
    assert clean_rows, f"www.{_BASE} not found in CSV"
    clean = clean_rows[0]
    assert clean["wildcard_attestation_status"] == "wildcard_not_detected"
    assert clean["wildcard_signature_matched"] == "false"
    assert clean["wildcard_differentiation_reason"] == "not_applicable"
    # Validate the promoted row
    prom_rows = [r for r in rows if r.get("tested_name") == f"mail.{_BASE}"]
    assert prom_rows, f"mail.{_BASE} not found in CSV"
    prom = prom_rows[0]
    assert prom["wildcard_attestation_status"] == "wildcard_detected"
    assert prom["wildcard_signature_matched"] == "false"
    assert prom["wildcard_differentiation_reason"] == REASON_DISTINCT_RRTYPE
    print("  PASS test_csv_s7_fields_correct (§10.8 CSV)")


def test_diagnostics_csv_s8_fields_correct() -> None:
    """§8 Diagnostics CSV: contains §8 columns with correct values for wildcard outcomes."""
    sup = _suppressed(f"mail.{_BASE}")
    wit = _withheld(f"news.{_BASE}")
    result = _make_run_result([], [sup, wit])
    out_dir = Path(tempfile.mkdtemp())
    diag_csv_path = export_diagnostics_csv(result, out_dir)
    with diag_csv_path.open(encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
    assert rows, "Diagnostics CSV must contain rows"
    required_cols = {"diagnostic_reason", "matched_rrtype", "matched_values"}
    assert reader.fieldnames is not None
    missing = required_cols - set(reader.fieldnames)
    assert not missing, f"Diagnostics CSV missing §8-only columns: {missing}"
    # Check suppressed row
    sup_rows = [r for r in rows if r.get("tested_name") == f"mail.{_BASE}"]
    assert sup_rows, f"mail.{_BASE} not found in diagnostics CSV"
    sup_row = sup_rows[0]
    assert sup_row["wildcard_attestation_status"] == "wildcard_detected"
    assert sup_row["wildcard_signature_matched"] == "true"
    assert sup_row["diagnostic_reason"] == "suppressed_as_wildcard_only"
    # Check withheld row
    wit_rows = [r for r in rows if r.get("tested_name") == f"news.{_BASE}"]
    assert wit_rows, f"news.{_BASE} not found in diagnostics CSV"
    wit_row = wit_rows[0]
    assert wit_row["wildcard_attestation_status"] == "wildcard_inconclusive"
    assert wit_row["wildcard_signature_matched"] == "not_applicable"
    assert wit_row["diagnostic_reason"] == "wildcard_attestation_inconclusive"
    print("  PASS test_diagnostics_csv_s8_fields_correct (§8 CSV)")


# ===========================================================================
# 7 — §10.8 XLSX: Findings and Diagnostics sheets
# ===========================================================================


def _xlsx_sheet_as_dicts(wb, sheet_name: str) -> list[dict[str, str]]:
    """Read an XLSX sheet as a list of dicts using the first row as header."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, [str(v) if v is not None else "" for v in row])) for row in rows[1:]]


def test_xlsx_findings_sheet_has_s7_columns() -> None:
    """§10.8 XLSX Findings sheet: all §7 column headers present."""
    records = [_confirmed_clean(f"www.{_BASE}")]
    result = _make_run_result(records, [])
    out_dir = Path(tempfile.mkdtemp())
    xlsx_path = export_xlsx_report(result, out_dir)
    wb = load_workbook(xlsx_path, read_only=True)
    assert "Findings" in wb.sheetnames, "XLSX must have a Findings sheet"
    ws = wb["Findings"]
    first_row = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
    expected = {
        "Wildcard attestation status",
        "Wildcard parent scope",
        "Wildcard probe count",
        "Wildcard RR types tested",
        "Wildcard signature matched",
        "Wildcard differentiation reason",
    }
    missing = expected - set(first_row)
    assert not missing, f"XLSX Findings sheet missing §7 headers: {missing}"
    print("  PASS test_xlsx_findings_sheet_has_s7_columns (§10.8 XLSX)")


def test_xlsx_diagnostics_sheet_has_s8_columns() -> None:
    """§8 XLSX Diagnostics sheet: §8-only columns present (matched_rrtype, diagnostic_reason)."""
    sup = _suppressed(f"mail.{_BASE}")
    result = _make_run_result([], [sup])
    out_dir = Path(tempfile.mkdtemp())
    xlsx_path = export_xlsx_report(result, out_dir)
    wb = load_workbook(xlsx_path, read_only=True)
    assert "Diagnostics" in wb.sheetnames, "XLSX must have a Diagnostics sheet"
    ws = wb["Diagnostics"]
    first_row = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
    expected_s8_headers = {"Matched RR type", "Matched values", "Diagnostic reason"}
    missing = expected_s8_headers - set(first_row)
    assert not missing, f"XLSX Diagnostics sheet missing §8 headers: {missing}"
    print("  PASS test_xlsx_diagnostics_sheet_has_s8_columns (§8 XLSX)")


def test_xlsx_suppressed_only_in_diagnostics_not_findings_sheet() -> None:
    """§10.7 XLSX: suppressed wildcard candidates appear on Diagnostics, not Findings sheet."""
    sup = _suppressed(f"mail.{_BASE}")
    clean_rec = _confirmed_clean(f"www.{_BASE}")
    result = _make_run_result([clean_rec], [sup])
    out_dir = Path(tempfile.mkdtemp())
    xlsx_path = export_xlsx_report(result, out_dir)
    wb = load_workbook(xlsx_path, read_only=True)
    findings_rows = _xlsx_sheet_as_dicts(wb, "Findings")
    diag_rows = _xlsx_sheet_as_dicts(wb, "Diagnostics")
    findings_names = {r.get("Tested name", "") for r in findings_rows}
    diag_names = {r.get("Tested name", "") for r in diag_rows}
    assert f"mail.{_BASE}" not in findings_names, (
        f"§10.7 XLSX: suppressed candidate must NOT be on Findings sheet"
    )
    assert f"mail.{_BASE}" in diag_names, (
        f"§10.7 XLSX: suppressed candidate must be on Diagnostics sheet"
    )
    # Confirm the §8 diagnostic_reason value on that row
    sup_diag = [r for r in diag_rows if r.get("Tested name") == f"mail.{_BASE}"]
    assert sup_diag, "Suppressed row must be found on Diagnostics sheet"
    assert sup_diag[0].get("Diagnostic reason") == "suppressed_as_wildcard_only", (
        f"§10.7 XLSX: diagnostic_reason must be suppressed_as_wildcard_only"
    )
    print("  PASS test_xlsx_suppressed_only_in_diagnostics_not_findings_sheet (§10.7 XLSX)")


# ===========================================================================
# 8 — JSON export: §7 fields on findings, §8 fields on evidence_diagnostics
# ===========================================================================


def test_json_s7_fields_on_confirmed_findings() -> None:
    """§10.8 JSON: confirmed findings include all §7 keys."""
    records = [
        _confirmed_clean(f"www.{_BASE}"),
        _confirmed_detected_promoted(f"mail.{_BASE}"),
    ]
    result = _make_run_result(records, [])
    out_dir = Path(tempfile.mkdtemp())
    json_path = export_json(result, out_dir)
    doc = json.loads(json_path.read_text(encoding="utf-8"))
    domain = doc["domains"][0]
    assert domain["findings"], "JSON must have at least one finding"
    required_keys = {
        "wildcard_attestation_status",
        "wildcard_parent_scope",
        "wildcard_probe_count",
        "wildcard_rrtypes_tested",
        "wildcard_signature_matched",
        "wildcard_differentiation_reason",
    }
    for finding in domain["findings"]:
        missing = required_keys - finding.keys()
        assert not missing, (
            f"JSON finding for {finding.get('tested_name')!r} missing §7 keys: {missing}"
        )
    print("  PASS test_json_s7_fields_on_confirmed_findings (§10.8 JSON)")


def test_json_s7_values_clean_finding() -> None:
    """JSON: clean finding carries wildcard_not_detected with not_applicable reason."""
    rec = _confirmed_clean(f"www.{_BASE}")
    result = _make_run_result([rec], [])
    out_dir = Path(tempfile.mkdtemp())
    json_path = export_json(result, out_dir)
    doc = json.loads(json_path.read_text(encoding="utf-8"))
    finding = doc["domains"][0]["findings"][0]
    assert finding["wildcard_attestation_status"] == "wildcard_not_detected"
    assert finding["wildcard_signature_matched"] == "false"
    assert finding["wildcard_differentiation_reason"] == "not_applicable"
    print("  PASS test_json_s7_values_clean_finding (JSON §7 valid combinations)")


def test_json_no_confirmed_finding_carries_wildcard_inconclusive() -> None:
    """Negative-action JSON: no finding in the JSON domains[].findings array carries wildcard_inconclusive."""
    records = [
        _confirmed_clean(f"www.{_BASE}"),
        _confirmed_detected_promoted(f"mail.{_BASE}"),
    ]
    result = _make_run_result(records, [])
    out_dir = Path(tempfile.mkdtemp())
    json_path = export_json(result, out_dir)
    doc = json.loads(json_path.read_text(encoding="utf-8"))
    for finding in doc["domains"][0]["findings"]:
        assert finding.get("wildcard_attestation_status") != "wildcard_inconclusive", (
            f"Negative-action JSON: confirmed finding carries wildcard_inconclusive"
        )
    print("  PASS test_json_no_confirmed_finding_carries_wildcard_inconclusive (JSON negative-action)")


def test_json_s8_fields_on_wildcard_diagnostics() -> None:
    """§8 JSON: evidence_diagnostics include wildcard §8 fields for suppressed/withheld outcomes."""
    sup = _suppressed(f"mail.{_BASE}")
    wit = _withheld(f"news.{_BASE}")
    result = _make_run_result([], [sup, wit])
    out_dir = Path(tempfile.mkdtemp())
    json_path = export_json(result, out_dir)
    doc = json.loads(json_path.read_text(encoding="utf-8"))
    diags = doc["domains"][0].get("evidence_diagnostics", [])
    assert diags, "JSON evidence_diagnostics must be non-empty"

    sup_diags = [d for d in diags if d.get("tested_name") == f"mail.{_BASE}"]
    assert sup_diags, f"mail.{_BASE} not found in JSON evidence_diagnostics"
    sup_diag = sup_diags[0]
    assert sup_diag.get("wildcard_attestation_status") == "wildcard_detected", (
        f"§8 JSON: suppressed -> wildcard_detected, got {sup_diag.get('wildcard_attestation_status')!r}"
    )
    assert sup_diag.get("wildcard_signature_matched") == "true", (
        f"§8 JSON: suppressed -> signature_matched=true, got {sup_diag.get('wildcard_signature_matched')!r}"
    )
    assert sup_diag.get("diagnostic_reason") == "suppressed_as_wildcard_only", (
        f"§8 JSON: suppressed -> diagnostic_reason=suppressed_as_wildcard_only"
    )

    wit_diags = [d for d in diags if d.get("tested_name") == f"news.{_BASE}"]
    assert wit_diags, f"news.{_BASE} not found in JSON evidence_diagnostics"
    wit_diag = wit_diags[0]
    assert wit_diag.get("wildcard_attestation_status") == "wildcard_inconclusive"
    assert wit_diag.get("wildcard_signature_matched") == "not_applicable"
    assert wit_diag.get("diagnostic_reason") == "wildcard_attestation_inconclusive"
    print("  PASS test_json_s8_fields_on_wildcard_diagnostics (§8 JSON)")


def test_json_s9_wildcard_note_field() -> None:
    """§9 JSON: wildcard_note field carries the contract phrasing for three cases."""
    rec_clean = _confirmed_clean(f"www.{_BASE}")
    rec_promoted = _confirmed_detected_promoted(f"mail.{_BASE}")
    sup = _suppressed(f"ftp.{_BASE}")
    wit = _withheld(f"smtp.{_BASE}")
    result = _make_run_result([rec_clean, rec_promoted], [sup, wit])
    out_dir = Path(tempfile.mkdtemp())
    json_path = export_json(result, out_dir)
    doc = json.loads(json_path.read_text(encoding="utf-8"))
    domain = doc["domains"][0]

    # Confirmed findings: promoted finding has the §9 phrasing; clean does not.
    findings_by_name = {f["tested_name"]: f for f in domain.get("findings", [])}
    promoted_finding = findings_by_name.get(f"mail.{_BASE}")
    assert promoted_finding is not None, f"mail.{_BASE} not found in JSON findings"
    assert promoted_finding.get("wildcard_note") == WILDCARD_WORDING_DETECTED_PROMOTED, (
        f"§9 JSON: promoted finding wildcard_note expected the detected-promoted phrasing"
    )
    clean_finding = findings_by_name.get(f"www.{_BASE}")
    assert clean_finding is not None
    assert clean_finding.get("wildcard_note", "") == "", (
        f"§9 JSON: clean finding must have empty wildcard_note"
    )

    # Diagnostics: suppressed has the suppressed phrasing; withheld has the inconclusive one.
    diags_by_name = {d["tested_name"]: d for d in domain.get("evidence_diagnostics", [])}
    sup_diag = diags_by_name.get(f"ftp.{_BASE}")
    assert sup_diag is not None, f"ftp.{_BASE} not found in JSON evidence_diagnostics"
    assert sup_diag.get("wildcard_note") == WILDCARD_WORDING_SUPPRESSED, (
        f"§9 JSON: suppressed wildcard_note must be the suppressed phrasing"
    )
    wit_diag = diags_by_name.get(f"smtp.{_BASE}")
    assert wit_diag is not None
    assert wit_diag.get("wildcard_note") == WILDCARD_WORDING_INCONCLUSIVE, (
        f"§9 JSON: withheld wildcard_note must be the inconclusive phrasing"
    )
    print("  PASS test_json_s9_wildcard_note_field (§9 JSON)")


# ===========================================================================
# 9 — Additional negative-action guards
# ===========================================================================


def test_not_detected_reason_is_not_applicable_not_a_differentiation_reason() -> None:
    """Negative-action: not_detected finding emits reason=not_applicable, not a differentiation reason."""
    differentiation_reasons = {
        "distinct_rrtype",
        "distinct_answer",
        "distinct_cname_target",
        "candidate_ns_soa",
        "verified_delegation",
    }
    rec = _confirmed_clean(f"www.{_BASE}")
    result = _make_run_result([rec], [])
    rows = build_confirmed_findings_rows(result)
    row = rows[0]
    reason = row["wildcard_differentiation_reason"]
    assert reason == "not_applicable", (
        f"not_detected finding must carry reason=not_applicable, got {reason!r}"
    )
    assert reason not in differentiation_reasons, (
        f"not_detected finding must NOT carry a differentiation reason; got {reason!r}"
    )
    print("  PASS test_not_detected_reason_is_not_applicable (negative-action)")


def test_reason_no_wildcard_maps_to_not_applicable() -> None:
    """REASON_NO_WILDCARD (='no_wildcard') maps to 'not_applicable' in the export layer."""
    # A record whose engine emitted REASON_NO_WILDCARD explicitly
    rec = DiscoveredRecord(
        fqdn=f"www.{_BASE}",
        record_type=RecordType.A,
        value="10.0.0.1",
        source_method="generated_candidate",
        classification=FindingClassification.STANDARD_RECORD,
        ttl=300,
        evidence_status=EvidenceStatus.CONFIRMED_ORDINARY_DNS_NAME,
        attestation_status="clean",
        wildcard_differentiation_reason=REASON_NO_WILDCARD,
    )
    result = _make_run_result([rec], [])
    rows = build_confirmed_findings_rows(result)
    row = rows[0]
    assert row["wildcard_differentiation_reason"] == "not_applicable", (
        f"REASON_NO_WILDCARD must map to not_applicable; got {row['wildcard_differentiation_reason']!r}"
    )
    print("  PASS test_reason_no_wildcard_maps_to_not_applicable (negative-action)")


# ===========================================================================
# Main
# ===========================================================================


def main() -> None:
    print("=== R4b Wildcard Reporting Contract Regression ===")
    test_prior_chain()

    print("\n--- §7 finding fields ---")
    test_s7_fields_present_on_clean_finding()
    test_s7_valid_combination_not_detected()
    test_s7_valid_combination_detected_promoted()
    test_s7_probe_count_and_rrtypes()
    test_s7_parent_scope_derived_from_fqdn()

    print("\n--- §7 negative-action: no confirmed finding carries wildcard_inconclusive ---")
    test_no_confirmed_finding_carries_wildcard_inconclusive()

    print("\n--- §8 diagnostic fields ---")
    test_s8_fields_on_suppressed_outcome()
    test_s8_fields_on_withheld_outcome()
    test_s8_parent_scope_on_diagnostic_row()
    test_s8_non_wildcard_outcome_has_blank_wildcard_fields()

    print("\n--- §9 report wording ---")
    test_s9_wording_detected_promoted_in_notes()
    test_s9_clean_finding_has_no_wildcard_detected_wording()
    test_s9_wording_suppressed_in_diagnostic_notes()
    test_s9_wording_inconclusive_in_diagnostic_notes()

    print("\n--- §10.7 sheet routing ---")
    test_s10_7_suppressed_only_in_diagnostics_not_findings()
    test_s10_7_withheld_only_in_diagnostics_not_findings()

    print("\n--- §10.8 CSV ---")
    test_csv_s7_fields_correct()
    test_diagnostics_csv_s8_fields_correct()

    print("\n--- §10.8 XLSX ---")
    test_xlsx_findings_sheet_has_s7_columns()
    test_xlsx_diagnostics_sheet_has_s8_columns()
    test_xlsx_suppressed_only_in_diagnostics_not_findings_sheet()

    print("\n--- JSON §7/§8/§9 ---")
    test_json_s7_fields_on_confirmed_findings()
    test_json_s7_values_clean_finding()
    test_json_no_confirmed_finding_carries_wildcard_inconclusive()
    test_json_s8_fields_on_wildcard_diagnostics()
    test_json_s9_wildcard_note_field()

    print("\n--- negative-action guards ---")
    test_not_detected_reason_is_not_applicable_not_a_differentiation_reason()
    test_reason_no_wildcard_maps_to_not_applicable()

    print("\n=== R4b: all assertions passed ===")


if __name__ == "__main__":
    main()
