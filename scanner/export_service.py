"""CSV, JSON, and XLSX export for DNS discovery scan results."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from scanner.paths import is_frozen
from scanner.input_loader import known_child_domains_from_record, normalize_domain_name
from scanner.input_loader import PREFERRED_INPUT_FORMAT_NOTE, RECOMMENDED_INPUT_COLUMNS_CSV
from scanner.models import (
    DiscoveredRecord,
    DomainInputRecord,
    DomainScanResult,
    FindingClassification,
    RecordType,
    ScanRunResult,
    ScanStatus,
)
from scanner.scan_engine import (
    CANDIDATE_STRONG_WARN_THRESHOLD,
    CANDIDATE_WARN_THRESHOLD,
    DNS_LIFETIME,
    DNS_TIMEOUT,
)

APP_NAME = ".US Locality DNS Discovery Tool"
APP_VERSION = None
EVIDENCE_MODEL_VERSION = "2.0-child-domain-discovery"
CHILD_DOMAIN_DISCOVERY_GOAL = (
    "For each known 3rd-level domain in the input, find child DNS names beneath it "
    "that are not already known in the system."
)
KNOWN_DOMAIN_DEFINITION = (
    "known_domain=yes when the discovered DNS name was already listed in the input "
    "known_fourth_level_domains or known_fifth_level_domains fields."
)
EVIDENCE_VALUE_DEFINITION = (
    "strong=new delegated child zone; moderate=new organizational/service child name; "
    "limited=new generic/technical hostname or base-zone-only; validation_only=known child confirmed; "
    "context_only=base zone exists without child discovery; none=no child names found; "
    "inconclusive=scan error or incomplete."
)
DOMAIN_DEFINITION_NOTE = (
    "In this report, domain means a DNS name beneath the scanned 3rd-level base domain. "
    "Some entries, such as www or autodiscover, may be service hostnames rather than "
    "separately registered domains."
)

OPERATOR_NOTE = (
    "This workbook supports unknown child domain discovery under known 3rd-level domains. "
    "It is not a complete zone inventory."
)

REVIEW_PATH_NOTE = (
    "Recommended review path: open Evidence Review first, focus on strong/moderate "
    "evidence_value rows with new_child_domains_found, use Verification guidance for "
    "optional independent dig checks, and rerun inconclusive rows before conclusions."
)

EVIDENCE_VALUE_ORDER = {
    "strong": 0,
    "moderate": 1,
    "limited": 2,
    "validation_only": 3,
    "context_only": 4,
    "inconclusive": 5,
    "none": 6,
}

RECOMMENDED_REVIEW_ACTIONS = {
    "strong": "Tool found delegated child zone — review NS/SOA evidence",
    "moderate": "Tool found new child DNS names — review organizational/service names",
    "limited": "Tool found generic/technical hostnames — limited value",
    "validation_only": "Tool confirmed known domains in DNS — informational only",
    "context_only": "Tool found base zone context only — no new child names",
    "inconclusive": "Rerun before conclusion",
    "none": "Tool found no new child domains by tested methods",
}

DISPLAY_RECOMMENDED_REVIEW_ACTIONS = {
    "strong": "Tool found NS/SOA evidence for a new delegated child zone",
    "moderate": "Tool found new child DNS names under this base domain",
    "limited": "Tool found only generic or technical hostnames",
    "validation_only": "Tool only confirmed domains already listed in the system input",
    "context_only": "Tool found base zone context only — no new child names",
    "inconclusive": "Rerun scan before drawing conclusions",
    "none": "Tool found no new child domains by tested methods",
}

GENERIC_HOSTNAME_LABELS = frozenset(
    {
        "www",
        "web",
        "ftp",
        "mail",
        "smtp",
        "imap",
        "pop",
        "pop3",
        "webmail",
        "vpn",
        "remote",
        "ns",
        "ns1",
        "ns2",
        "mx",
        "mx1",
        "mx2",
    }
)

TECHNICAL_VENDOR_LABELS = frozenset(
    {
        "autodiscover",
        "msoid",
        "lyncdiscover",
        "sip",
        "enterpriseregistration",
        "enterpriseenrollment",
    }
)

ORGANIZATIONAL_LABELS = frozenset(
    {
        "police",
        "fire",
        "library",
        "court",
        "clerk",
        "records",
        "gis",
        "tax",
        "assessor",
        "water",
        "sewer",
        "publicworks",
        "planning",
        "permits",
        "finance",
        "cityhall",
        "townhall",
        "parks",
        "admin",
        "ci",
        "co",
    }
)

SERVICE_HOSTNAME_LABELS = frozenset(
    {
        "portal",
        "services",
        "service",
        "owa",
        "citrix",
    }
)

EVIDENCE_REVIEW_COLUMNS = [
    "base_domain",
    "delegated_manager",
    "evidence_value",
    "recommended_review_action",
    "new_child_domains_found",
    "new_child_domains_count",
    "new_delegated_domains_found",
    "new_organizational_domains_found",
    "new_generic_hostnames_found",
    "new_technical_hostnames_found",
    "known_domains_validated",
    "why",
    "manual_verification_hint",
    "limitation_note",
]

DISCOVERY_LIMITATION = (
    "Child domain discovery results show only DNS names found through tested methods. "
    "No child names discovered does not prove that no child domains exist."
)

CONTEXT_LIMITATION = (
    "Results compare known domains from the system input against child DNS names found "
    "by live testing. The scan does not enumerate every possible child domain. "
    "A few defensible new examples may be enough for the business question."
)

PARTIAL_SCAN_NOTE = (
    "This scan was cancelled before all domains were completed. Results are partial."
)

SCAN_ERROR_NOTE = "Domain status is incomplete/error; rerun recommended."

SOA_FINDING_NOTE = (
    "SOA discovered; zone exists even though requested record type may have no direct answer."
)
FIFTH_LEVEL_PARENT_NOTE = (
    "Parent checked because 5th-level candidates were tested under this name."
)

CSV_COLUMNS = [
    "scan_timestamp",
    "base_domain",
    "discovered_name",
    "known_domain",
    "name_type",
    "evidence_value",
    "why",
    "tested_name",
    "record_type",
    "finding_type",
    "confidence",
    "source",
    "nameserver",
    "value",
    "ttl",
    "wildcard_suspected",
    "axfr_status",
    "error",
    "wordlist_sources",
    "notes",
]

SUMMARY_COLUMNS = [
    "scan_timestamp",
    "base_domain",
    "input_domain",
    "delegated_manager",
    "zone",
    "second_level_domain",
    "known_domains_from_system",
    "known_domains_from_system_count",
    "known_domains_validated",
    "known_domains_validated_count",
    "new_child_domains_found",
    "new_child_domains_count",
    "new_delegated_domains_found",
    "new_delegated_domains_count",
    "new_generic_hostnames_found",
    "new_generic_hostnames_count",
    "new_technical_hostnames_found",
    "new_technical_hostnames_count",
    "new_organizational_domains_found",
    "new_organizational_domains_count",
    "evidence_value",
    "why",
    "recommended_review_action",
    "manual_verification_hint",
    "scan_status",
    "authoritative_nameservers",
    "axfr_status",
    "wildcard_suspected",
    "base_zone_exists",
    "candidate_names_tested",
    "wordlist_sources",
    "analysis_note",
    "limitation_note",
]

ERRORS_WARNINGS_COLUMNS = [
    "scan_timestamp",
    "base_domain",
    "delegated_manager",
    "zone",
    "warning_type",
    "tested_name",
    "record_type",
    "nameserver",
    "message",
    "notes",
]

SCAN_STATUS_AXFR = "AXFR allowed"
SCAN_STATUS_DELEGATED_CHILD = "Possible delegated child zone discovered"
SCAN_STATUS_DNS_ACTIVITY_WITH_ERRORS = "DNS activity discovered with scan errors"
SCAN_STATUS_DNS_ACTIVITY = "DNS activity discovered"
SCAN_STATUS_BASE_ZONE_EXISTS = "Base domain zone exists"
SCAN_STATUS_BASE_ONLY = "Base domain records only"
SCAN_STATUS_INCOMPLETE = "Scan incomplete / error"
SCAN_STATUS_ERRORS_ONLY = "Scan errors only"
SCAN_STATUS_NO_RECORDS = "No records discovered using tested methods"

# Legacy alias used in a few internal references.
SCAN_STATUS_SUBDELEGATION = SCAN_STATUS_DELEGATED_CHILD

SUMMARY_STATUS_FILLS = {
    SCAN_STATUS_AXFR: PatternFill(fill_type="solid", fgColor="C6EFCE"),
    SCAN_STATUS_DELEGATED_CHILD: PatternFill(fill_type="solid", fgColor="BDD7EE"),
    SCAN_STATUS_DNS_ACTIVITY_WITH_ERRORS: PatternFill(fill_type="solid", fgColor="F4B084"),
    SCAN_STATUS_DNS_ACTIVITY: PatternFill(fill_type="solid", fgColor="FFE699"),
    SCAN_STATUS_BASE_ZONE_EXISTS: PatternFill(fill_type="solid", fgColor="D9E1F2"),
    SCAN_STATUS_BASE_ONLY: PatternFill(fill_type="solid", fgColor="EDEDED"),
    SCAN_STATUS_INCOMPLETE: PatternFill(fill_type="solid", fgColor="FFC7CE"),
    SCAN_STATUS_ERRORS_ONLY: PatternFill(fill_type="solid", fgColor="FFC7CE"),
    SCAN_STATUS_NO_RECORDS: PatternFill(fill_type="solid", fgColor="F8CBAD"),
}

EVIDENCE_VALUE_FILLS = {
    "Strong": PatternFill(fill_type="solid", fgColor="C6EFCE"),
    "Moderate": PatternFill(fill_type="solid", fgColor="BDD7EE"),
    "Limited": PatternFill(fill_type="solid", fgColor="FFE699"),
    "Validation only": PatternFill(fill_type="solid", fgColor="D9E1F2"),
    "Context only": PatternFill(fill_type="solid", fgColor="EDEDED"),
    "None": PatternFill(fill_type="solid", fgColor="F8CBAD"),
    "Inconclusive": PatternFill(fill_type="solid", fgColor="FFC7CE"),
}


@dataclass(frozen=True)
class SheetColumn:
    """Internal row key plus coworker-facing XLSX header label."""

    key: str
    label: str


def _sheet_columns(keys: list[str], labels: dict[str, str]) -> list[SheetColumn]:
    return [SheetColumn(key, labels.get(key, key.replace("_", " ").title())) for key in keys]


EVIDENCE_REVIEW_HEADER_LABELS = {
    "base_domain": "Base domain",
    "delegated_manager": "Delegated manager",
    "evidence_value": "Evidence value",
    "recommended_review_action": "Recommended review action",
    "new_child_domains_found": "New child domains found",
    "new_child_domains_count": "New child domains count",
    "new_delegated_domains_found": "New delegated domains found",
    "new_organizational_domains_found": "New organizational domains found",
    "new_generic_hostnames_found": "New generic hostnames found",
    "new_technical_hostnames_found": "New technical hostnames found",
    "known_domains_validated": "Known domains validated",
    "why": "Why",
    "manual_verification_hint": "Verification guidance",
    "limitation_note": "Limitation note",
}

EVIDENCE_REVIEW_SHEET_COLUMNS = _sheet_columns(
    EVIDENCE_REVIEW_COLUMNS,
    EVIDENCE_REVIEW_HEADER_LABELS,
)

SUMMARY_HEADER_LABELS = {
    "scan_timestamp": "Scan timestamp",
    "base_domain": "Base domain",
    "input_domain": "Input domain",
    "delegated_manager": "Delegated manager",
    "zone": "Zone",
    "second_level_domain": "Second level domain",
    "known_domains_from_system": "Known domains from system",
    "known_domains_from_system_count": "Known domains from system count",
    "known_domains_validated": "Known domains validated",
    "known_domains_validated_count": "Known domains validated count",
    "new_child_domains_found": "New child domains found",
    "new_child_domains_count": "New child domains count",
    "new_delegated_domains_found": "New delegated domains found",
    "new_delegated_domains_count": "New delegated domains count",
    "new_generic_hostnames_found": "New generic hostnames found",
    "new_generic_hostnames_count": "New generic hostnames count",
    "new_technical_hostnames_found": "New technical hostnames found",
    "new_technical_hostnames_count": "New technical hostnames count",
    "new_organizational_domains_found": "New organizational domains found",
    "new_organizational_domains_count": "New organizational domains count",
    "evidence_value": "Evidence value",
    "why": "Why",
    "recommended_review_action": "Recommended review action",
    "manual_verification_hint": "Verification guidance",
    "scan_status": "Scan status",
    "authoritative_nameservers": "Authoritative nameservers",
    "axfr_status": "AXFR status",
    "wildcard_suspected": "Wildcard suspected",
    "base_zone_exists": "Base zone exists",
    "candidate_names_tested": "Candidate names tested",
    "wordlist_sources": "Wordlist sources",
    "analysis_note": "Analysis note",
    "limitation_note": "Limitation note",
}

SUMMARY_SHEET_COLUMNS = _sheet_columns(SUMMARY_COLUMNS, SUMMARY_HEADER_LABELS)

FINDINGS_XLSX_COLUMN_KEYS = [
    "discovered_name",
    "known_domain",
    "name_type",
    "evidence_value",
    "why",
    "base_domain",
    "tested_name",
    "record_type",
    "finding_type",
    "confidence",
    "source",
    "nameserver",
    "value",
    "ttl",
    "wildcard_suspected",
    "axfr_status",
    "error",
    "wordlist_sources",
    "notes",
    "scan_timestamp",
]

FINDINGS_HEADER_LABELS = {
    "scan_timestamp": "Scan timestamp",
    "base_domain": "Base domain",
    "discovered_name": "Discovered name",
    "known_domain": "Known domain",
    "name_type": "Name type",
    "evidence_value": "Evidence value",
    "why": "Why",
    "tested_name": "Tested name",
    "record_type": "Record type",
    "finding_type": "Finding type",
    "confidence": "Confidence",
    "source": "Source",
    "nameserver": "Nameserver",
    "value": "Value",
    "ttl": "TTL",
    "wildcard_suspected": "Wildcard suspected",
    "axfr_status": "AXFR status",
    "error": "Error",
    "wordlist_sources": "Wordlist sources",
    "notes": "Notes",
}

FINDINGS_SHEET_COLUMNS = _sheet_columns(FINDINGS_XLSX_COLUMN_KEYS, FINDINGS_HEADER_LABELS)

ERRORS_WARNINGS_HEADER_LABELS = {
    "scan_timestamp": "Scan timestamp",
    "base_domain": "Base domain",
    "delegated_manager": "Delegated manager",
    "zone": "Zone",
    "warning_type": "Warning type",
    "tested_name": "Tested name",
    "record_type": "Record type",
    "nameserver": "Nameserver",
    "message": "Message",
    "notes": "Notes",
}

ERRORS_WARNINGS_SHEET_COLUMNS = _sheet_columns(
    ERRORS_WARNINGS_COLUMNS,
    ERRORS_WARNINGS_HEADER_LABELS,
)

ExportFormat = Literal["xlsx", "csv", "json", "all"]


@dataclass
class ExportOutcome:
    """Result of an export operation."""

    xlsx_path: Path | None = None
    csv_path: Path | None = None
    json_path: Path | None = None
    summary_csv_path: Path | None = None
    row_count: int = 0
    domain_count: int = 0


def _format_timestamp(value: datetime | None) -> str:
    if value is None:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return value.strftime("%Y-%m-%d %H:%M:%S")


def _findings_report_stem(scan_timestamp: datetime | None) -> str:
    stamp = scan_timestamp or datetime.now()
    return f"us_locality_dns_discovery_{stamp.strftime('%Y%m%d_%H%M%S')}"


def _workbook_report_stem(scan_timestamp: datetime | None) -> str:
    stamp = scan_timestamp or datetime.now()
    return f"us_locality_dns_report_{stamp.strftime('%Y%m%d_%H%M%S')}"


def _export_notes(result: ScanRunResult) -> str:
    parts = [DISCOVERY_LIMITATION, CONTEXT_LIMITATION]
    if result.partial or result.cancelled:
        parts.append(PARTIAL_SCAN_NOTE)
    return " ".join(parts)


def _limitation_note(result: ScanRunResult) -> str:
    return _export_notes(result)


def _input_metadata_values(domain_result: DomainScanResult) -> dict[str, str]:
    record = domain_result.input_record
    if record is None:
        return {
            "input_domain": "",
            "delegated_manager": "",
            "zone": "",
            "second_level_domain": "",
            "known_fourth_level_count": "",
            "known_fifth_level_count": "",
            "known_fourth_level_domains": "",
            "known_fifth_level_domains": "",
        }

    return {
        "input_domain": record.original_domain,
        "delegated_manager": record.delegated_manager,
        "zone": record.zone,
        "second_level_domain": record.second_level_domain,
        "known_fourth_level_count": record.fourth_level_count,
        "known_fifth_level_count": record.fifth_level_count,
        "known_fourth_level_domains": "; ".join(record.known_fourth_level_domains),
        "known_fifth_level_domains": "; ".join(record.known_fifth_level_domains),
    }


def _is_child_name(fqdn: str, base_domain: str) -> bool:
    child = normalize_domain_name(fqdn)
    base = normalize_domain_name(base_domain)
    if child == base:
        return False
    return child.endswith(f".{base}")


def _join_domain_list(names: set[str]) -> str:
    return "; ".join(sorted(names))


DISPLAY_KNOWN_DOMAIN = {
    "yes": "Yes",
    "no": "No",
}

DISPLAY_EVIDENCE_VALUE = {
    "strong": "Strong",
    "moderate": "Moderate",
    "limited": "Limited",
    "validation_only": "Validation only",
    "context_only": "Context only",
    "none": "None",
    "inconclusive": "Inconclusive",
}

DISPLAY_NAME_TYPE = {
    "base_domain": "Base domain",
    "delegated_child_zone": "Delegated child zone",
    "organizational_child_name": "Organizational child name",
    "service_hostname": "Service hostname",
    "generic_hostname": "Generic hostname",
    "technical_vendor_hostname": "Technical/vendor hostname",
    "alias_to_known_domain": "Alias to known domain",
    "alias_to_external_target": "Alias to external target",
    "unknown_child_hostname": "Unknown child hostname",
}


@dataclass
class DnsEvidenceContext:
    """DNS evidence indexed by owner name for CNAME-aware classification."""

    cname_targets: dict[str, str] = field(default_factory=dict)
    direct_delegation: set[str] = field(default_factory=set)


def _display_known_domain(value: str) -> str:
    return DISPLAY_KNOWN_DOMAIN.get(value.lower(), value)


def _display_evidence_value(value: str) -> str:
    return DISPLAY_EVIDENCE_VALUE.get(value.lower(), value)


def _display_name_type(value: str) -> str:
    return DISPLAY_NAME_TYPE.get(value, value.replace("_", " ").title())


def _display_boolean(value: str) -> str:
    lowered = value.strip().lower()
    if lowered in {"true", "yes", "1"}:
        return "Yes"
    if lowered in {"false", "no", "0"}:
        return "No"
    return value


def _coworker_display_row(row: dict[str, str]) -> dict[str, str]:
    """Return a copy with human-readable classification fields for XLSX."""
    display = dict(row)
    evidence_code = display.get("evidence_value", "").lower()
    if evidence_code in DISPLAY_RECOMMENDED_REVIEW_ACTIONS:
        display["recommended_review_action"] = DISPLAY_RECOMMENDED_REVIEW_ACTIONS[evidence_code]
    if "known_domain" in display:
        display["known_domain"] = _display_known_domain(display["known_domain"])
    if "evidence_value" in display:
        display["evidence_value"] = _display_evidence_value(display["evidence_value"])
    if "name_type" in display and display["name_type"]:
        display["name_type"] = _display_name_type(display["name_type"])
    for bool_field in ("wildcard_suspected", "base_zone_exists"):
        if bool_field in display and display[bool_field]:
            display[bool_field] = _display_boolean(display[bool_field])
    return display


def _build_dns_evidence_context(domain_result: DomainScanResult) -> DnsEvidenceContext:
    """Build CNAME and direct-delegation maps from raw DNS findings."""
    base = domain_result.domain
    cname_targets: dict[str, str] = {}
    direct_delegation: set[str] = set()

    for record in domain_result.records:
        fqdn = normalize_domain_name(record.fqdn)

        if record.record_type == RecordType.CNAME:
            target = normalize_domain_name(record.value.rstrip("."))
            cname_targets[fqdn] = target

        if record.classification == FindingClassification.DELEGATED_CHILD_ZONE:
            if record.record_type == RecordType.NS and _is_child_name(fqdn, base):
                direct_delegation.add(fqdn)
            continue

        if record.classification == FindingClassification.ZONE_SOA_DISCOVERED:
            if _is_child_name(fqdn, base):
                direct_delegation.add(fqdn)
            continue

        if record.classification == FindingClassification.AXFR_SUCCESS:
            if _is_child_name(fqdn, base):
                direct_delegation.add(fqdn)

    return DnsEvidenceContext(cname_targets=cname_targets, direct_delegation=direct_delegation)


def _has_direct_delegation(fqdn: str, ctx: DnsEvidenceContext) -> bool:
    return normalize_domain_name(fqdn) in ctx.direct_delegation


def _cname_target_for(fqdn: str, ctx: DnsEvidenceContext) -> str | None:
    return ctx.cname_targets.get(normalize_domain_name(fqdn))


def _why_for_child_name(
    *,
    fqdn: str,
    base_domain: str,
    known_domain: bool,
    name_type: str,
    evidence_value: str,
    ctx: DnsEvidenceContext,
    record: DiscoveredRecord | None = None,
) -> str:
    if record is not None and record.source_method == "fifth_level_parent_validation":
        if known_domain:
            return (
                "Tool confirmed a 4th-level domain already listed in the system input "
                "while checking deeper candidate names."
            )
        return (
            "Tool found direct DNS evidence for this 4th-level child domain while checking "
            "deeper candidate names."
        )

    if name_type == "base_domain":
        if evidence_value == "context_only":
            return "Only base-domain DNS evidence was found; no child domains were discovered by tested methods."
        return "Base domain DNS records discovered during the scan."

    if name_type == "alias_to_known_domain":
        return (
            "New DNS alias/hostname found, but it points to a child domain already known in the system."
        )

    if name_type == "alias_to_external_target":
        return (
            "External CNAME target is not counted as a child domain under the scanned base domain."
        )

    if known_domain and name_type == "delegated_child_zone":
        return "Already listed in the system and confirmed in DNS."

    if known_domain:
        return "Already listed in the system and confirmed in DNS."

    if name_type == "delegated_child_zone":
        return "New delegated child domain found in live DNS and not listed in the system input."

    if name_type == "generic_hostname":
        return (
            "New hostname found in live DNS, but it appears to be a generic or technical service name."
        )

    if name_type == "technical_vendor_hostname":
        return (
            "New hostname found in live DNS, but it appears to be a generic or technical service name."
        )

    if name_type in {"organizational_child_name", "service_hostname"}:
        return "New child domain found in live DNS and not listed in the system input."

    if evidence_value == "limited":
        target = _cname_target_for(fqdn, ctx)
        if target and _is_child_name(target, base_domain):
            return (
                "New DNS alias/hostname found, but it points to a child domain already known "
                "in the system."
            )
        return "New child domain found in live DNS and not listed in the system input."

    return "New child domain found in live DNS and not listed in the system input."


def _why_for_domain_summary(
    evidence_value: str,
    comparison: dict[str, str | int | bool],
    *,
    scan_failed: bool,
) -> str:
    if scan_failed or evidence_value == "inconclusive":
        return "Scan was incomplete or encountered errors. Rerun before drawing conclusions."

    new_count = int(comparison.get("new_child_domains_count", 0))
    new_delegated_count = int(comparison.get("new_delegated_domains_count", 0))
    validated_count = int(comparison.get("known_domains_validated_count", 0))

    if evidence_value == "strong" or new_delegated_count > 0:
        return (
            "Tool found NS/SOA evidence for a new delegated child zone not listed "
            "in the system input."
        )

    if evidence_value == "moderate" and new_count > 0:
        return "Tool found new child DNS names under this base domain not listed in the system input."

    if evidence_value == "limited" and new_count > 0:
        if new_delegated_count == 0 and validated_count > 0:
            return (
                "Tool found a DNS alias/hostname, but it points to a child domain already "
                "known in the system."
            )
        return (
            "Tool found hostnames in live DNS, but they appear to be generic or technical "
            "service names."
        )

    if evidence_value == "validation_only" or (validated_count > 0 and new_count == 0):
        return "Tool only confirmed domains already listed in the system input."

    if evidence_value == "context_only":
        return (
            "Tool found only base-domain DNS evidence; no child domains were discovered "
            "by tested methods."
        )

    if evidence_value == "none":
        return "Tool did not find child DNS names with the selected methods."

    return "Tool did not find child DNS names with the selected methods."


def _collect_dns_discovered_children(
    domain_result: DomainScanResult,
) -> tuple[set[str], set[str], bool]:
    """
    Return DNS-discovered child names, delegated child zone names, and base-only flag.

    Child names come from standard_record, delegated_child_zone, zone_soa_discovered
    (below base), and axfr_success (below base).
    """
    base = domain_result.domain
    child_names: set[str] = set()
    delegated_children: set[str] = set()
    base_evidence = False

    for record in domain_result.records:
        classification = record.classification
        fqdn = normalize_domain_name(record.fqdn)

        if classification in {
            FindingClassification.QUERY_ERROR,
            FindingClassification.SCAN_ERROR,
            FindingClassification.AXFR_BLOCKED,
            FindingClassification.NO_RECORDS_DISCOVERED,
            FindingClassification.AUTHORITATIVE_NS,
        }:
            continue

        if classification in {
            FindingClassification.BASE_DOMAIN_RECORD,
            FindingClassification.BASE_ZONE_EXISTS,
        }:
            if fqdn == normalize_domain_name(base):
                base_evidence = True
            continue

        if classification == FindingClassification.ZONE_SOA_DISCOVERED:
            if _is_child_name(fqdn, base):
                child_names.add(fqdn)
            elif fqdn == normalize_domain_name(base):
                base_evidence = True
            continue

        if classification == FindingClassification.DELEGATED_CHILD_ZONE:
            if _is_child_name(fqdn, base):
                child_names.add(fqdn)
                delegated_children.add(fqdn)
            continue

        if classification == FindingClassification.STANDARD_RECORD:
            if _is_child_name(fqdn, base):
                child_names.add(fqdn)
            continue

        if classification == FindingClassification.AXFR_SUCCESS:
            if _is_child_name(fqdn, base):
                child_names.add(fqdn)
            elif fqdn == normalize_domain_name(base):
                base_evidence = True
            continue

    dns_discovered_base_only = base_evidence and not child_names
    return child_names, delegated_children, dns_discovered_base_only


def _extract_leftmost_label(fqdn: str, base_domain: str) -> str:
    base = normalize_domain_name(base_domain)
    child = normalize_domain_name(fqdn)
    if child == base or not child.endswith(f".{base}"):
        return ""
    relative = child[: -(len(base) + 1)]
    return relative.split(".")[0] if relative else ""


def _classify_name_type(
    fqdn: str,
    base_domain: str,
    *,
    is_delegated: bool,
    ctx: DnsEvidenceContext | None = None,
    known: set[str] | None = None,
) -> str:
    normalized = normalize_domain_name(fqdn)
    if normalized == normalize_domain_name(base_domain):
        return "base_domain"

    if ctx is not None:
        target = _cname_target_for(fqdn, ctx)
        if target is not None:
            if not _is_child_name(target, base_domain):
                return "alias_to_external_target"
            if known and target in known:
                return "alias_to_known_domain"

    if is_delegated:
        return "delegated_child_zone"
    label = _extract_leftmost_label(fqdn, base_domain)
    if label in GENERIC_HOSTNAME_LABELS:
        return "generic_hostname"
    if label in TECHNICAL_VENDOR_LABELS:
        return "technical_vendor_hostname"
    if label in ORGANIZATIONAL_LABELS:
        return "organizational_child_name"
    if label in SERVICE_HOSTNAME_LABELS:
        return "service_hostname"
    return "unknown_child_hostname"


def _evidence_value_for_child_name(
    *,
    known_domain: bool,
    name_type: str,
    is_delegated: bool,
    has_axfr_child: bool,
) -> str:
    if known_domain:
        return "validation_only"
    if name_type in {"alias_to_known_domain", "alias_to_external_target"}:
        return "limited"
    if is_delegated or name_type == "delegated_child_zone" or has_axfr_child:
        return "strong"
    if name_type in {"organizational_child_name", "service_hostname"}:
        return "moderate"
    if name_type in {"generic_hostname", "technical_vendor_hostname", "unknown_child_hostname"}:
        return "limited"
    return "limited"


def _aggregate_domain_evidence_value(
    child_evidence_values: list[str],
    *,
    scan_failed: bool,
    has_base_zone: bool,
    any_child_discovered: bool,
    any_known_validated: bool,
) -> str:
    if scan_failed:
        return "inconclusive"
    new_values = [
        value
        for value in child_evidence_values
        if value in {"strong", "moderate", "limited"}
    ]
    if "strong" in new_values:
        return "strong"
    if "moderate" in new_values:
        return "moderate"
    if "limited" in new_values:
        return "limited"
    if any_known_validated:
        return "validation_only"
    if has_base_zone and not any_child_discovered:
        return "context_only"
    if not any_child_discovered:
        return "none"
    return "none"


def _classify_child_name_metadata(
    name: str,
    domain_result: DomainScanResult,
    known: set[str],
    ctx: DnsEvidenceContext,
    axfr_new: set[str],
) -> dict[str, str]:
    is_delegated = _has_direct_delegation(name, ctx)
    known_domain = name in known
    name_type = _classify_name_type(
        name,
        domain_result.domain,
        is_delegated=is_delegated,
        ctx=ctx,
        known=known,
    )
    evidence_value = _evidence_value_for_child_name(
        known_domain=known_domain,
        name_type=name_type,
        is_delegated=is_delegated,
        has_axfr_child=name in axfr_new,
    )
    why = _why_for_child_name(
        fqdn=name,
        base_domain=domain_result.domain,
        known_domain=known_domain,
        name_type=name_type,
        evidence_value=evidence_value,
        ctx=ctx,
    )
    return {
        "discovered_name": name,
        "known_domain": "yes" if known_domain else "no",
        "name_type": name_type,
        "evidence_value": evidence_value,
        "why": why,
    }


def _build_child_domain_inventory(
    domain_result: DomainScanResult,
) -> dict[str, str | int | bool | set[str]]:
    known = known_child_domains_from_record(domain_result.input_record)
    ctx = _build_dns_evidence_context(domain_result)
    discovered, delegated_discovered, dns_discovered_base_only = _collect_dns_discovered_children(
        domain_result
    )
    axfr_new = _axfr_child_names_not_in_input(domain_result, known)

    per_name: dict[str, dict[str, str]] = {}
    for name in sorted(discovered | axfr_new):
        per_name[name] = _classify_child_name_metadata(
            name,
            domain_result,
            known,
            ctx,
            axfr_new,
        )

    validated = discovered & known
    new_children = {name for name, meta in per_name.items() if meta["known_domain"] == "no"}
    new_delegated = {
        name
        for name, meta in per_name.items()
        if meta["known_domain"] == "no" and meta["name_type"] == "delegated_child_zone"
    }
    new_generic = {
        name
        for name, meta in per_name.items()
        if meta["known_domain"] == "no"
        and meta["name_type"] in {"generic_hostname", "alias_to_known_domain"}
    }
    new_technical = {
        name
        for name, meta in per_name.items()
        if meta["known_domain"] == "no" and meta["name_type"] == "technical_vendor_hostname"
    }
    new_organizational = {
        name
        for name, meta in per_name.items()
        if meta["known_domain"] == "no"
        and meta["name_type"] in {"organizational_child_name", "service_hostname"}
    }

    counts = _domain_summary_counts(domain_result)
    unknown_evidence = [
        meta["evidence_value"]
        for meta in per_name.values()
        if meta["known_domain"] == "no"
    ]
    evidence_value = _aggregate_domain_evidence_value(
        unknown_evidence,
        scan_failed=_domain_has_scan_error(domain_result),
        has_base_zone=bool(counts["base_zone_exists_flag"]),
        any_child_discovered=bool(discovered),
        any_known_validated=bool(validated),
    )

    return {
        "known_domains_from_system": _join_domain_list(known),
        "known_domains_from_system_count": len(known),
        "known_domains_validated": _join_domain_list(validated),
        "known_domains_validated_count": len(validated),
        "new_child_domains_found": _join_domain_list(new_children),
        "new_child_domains_count": len(new_children),
        "new_delegated_domains_found": _join_domain_list(new_delegated),
        "new_delegated_domains_count": len(new_delegated),
        "new_generic_hostnames_found": _join_domain_list(new_generic),
        "new_generic_hostnames_count": len(new_generic),
        "new_technical_hostnames_found": _join_domain_list(new_technical),
        "new_technical_hostnames_count": len(new_technical),
        "new_organizational_domains_found": _join_domain_list(new_organizational),
        "new_organizational_domains_count": len(new_organizational),
        "evidence_value": evidence_value,
        "_per_name": per_name,
        "_new_children_set": new_children,
        "_new_delegated_set": new_delegated,
        "_dns_discovered_base_only": dns_discovered_base_only,
        "_known_set": known,
        "_dns_ctx": ctx,
    }


def _compare_known_vs_discovered(
    domain_result: DomainScanResult,
) -> dict[str, str | int | bool]:
    inventory = _build_child_domain_inventory(domain_result)
    export = {key: value for key, value in inventory.items() if not str(key).startswith("_")}
    export["_new_children_set"] = inventory["_new_children_set"]
    export["_new_delegated_set"] = inventory["_new_delegated_set"]
    export["_known_set"] = inventory["_known_set"]
    export["_dns_discovered_base_only"] = inventory["_dns_discovered_base_only"]
    export["_per_name"] = inventory["_per_name"]
    export["_dns_ctx"] = inventory["_dns_ctx"]
    return export


def _axfr_child_names_not_in_input(domain_result: DomainScanResult, known: set[str]) -> set[str]:
    base = normalize_domain_name(domain_result.domain)
    not_in_input: set[str] = set()
    for record in domain_result.records:
        if record.classification != FindingClassification.AXFR_SUCCESS:
            continue
        fqdn = normalize_domain_name(record.fqdn)
        if fqdn != base and _is_child_name(fqdn, domain_result.domain) and fqdn not in known:
            not_in_input.add(fqdn)
    return not_in_input


def _evidence_support_level(
    domain_result: DomainScanResult,
    counts: dict[str, int],
    scan_status: str,
    comparison: dict[str, str | int | bool],
) -> str:
    """Legacy alias — returns evidence_value."""
    return str(comparison.get("evidence_value", "none"))


def _analysis_note(
    domain_result: DomainScanResult,
    counts: dict[str, int],
    scan_status: str,
    comparison: dict[str, str | int | bool],
) -> str:
    if _domain_has_scan_error(domain_result):
        return "Scan was incomplete or encountered errors. Rerun before drawing conclusions."

    new_count = int(comparison.get("new_child_domains_count", 0))
    validated_count = int(comparison.get("known_domains_validated_count", 0))
    known_count = int(comparison.get("known_domains_from_system_count", 0))
    dns_discovered_base_only = bool(comparison.get("_dns_discovered_base_only"))

    if new_count > 0:
        return (
            f"Tool found {new_count} child DNS name(s) in live DNS not already listed "
            "in the system known-domain fields."
        )

    if validated_count > 0 and known_count > 0:
        return (
            "Tool confirmed known child domains from the system input in live DNS; "
            "no new child names found."
        )

    if known_count > 0 and validated_count == 0:
        return (
            "Input lists known child domains, but this scan did not validate them in live DNS."
        )

    if dns_discovered_base_only:
        return (
            "Tool found base zone evidence, but no child DNS names were discovered "
            "by tested methods."
        )

    if scan_status == SCAN_STATUS_NO_RECORDS:
        return (
            "Tool did not find child DNS names with the selected methods; "
            "this does not prove absence."
        )

    return "Tool found no new child domains using tested methods."


def _recommended_review_action(evidence_value: str) -> str:
    return RECOMMENDED_REVIEW_ACTIONS.get(
        evidence_value,
        "No action / no evidence found",
    )


def _parse_domain_list(value: str) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in value.split(";") if part.strip()]


def _manual_verification_hint(
    domain_result: DomainScanResult,
    counts: dict[str, int],
    scan_status: str,
    comparison: dict[str, str | int | bool],
    evidence_value: str,
) -> str:
    if evidence_value == "inconclusive" or _domain_has_scan_error(domain_result):
        return "Scan was incomplete or encountered errors. Rerun before drawing conclusions."

    if evidence_value == "none":
        return (
            "Tool did not find child DNS names with the selected methods. "
            "No manual check is suggested unless this domain is important."
        )

    base = domain_result.domain
    per_name = comparison.get("_per_name", {})
    new_delegated = comparison.get("_new_delegated_set", set())
    new_children = comparison.get("_new_children_set", set())

    if evidence_value == "strong" and new_delegated:
        child = sorted(new_delegated)[0]
        return (
            f"Tool found NS/SOA evidence for {child}, indicating a delegated child zone. "
            f"Optional manual check: dig NS {child}; dig SOA {child}"
        )

    if evidence_value == "moderate" and new_children:
        for name in sorted(new_children):
            meta = per_name.get(name, {}) if isinstance(per_name, dict) else {}
            if meta.get("name_type") in {"organizational_child_name", "service_hostname"}:
                return (
                    f"Tool found DNS records for {name}. "
                    f"Optional manual check: dig A {name}; dig NS {name}"
                )
        child = sorted(new_children)[0]
        return (
            f"Tool found DNS records for {child}. "
            f"Optional manual check: dig A {child}; dig CNAME {child}"
        )

    if evidence_value == "limited" and new_children:
        for name in sorted(new_children):
            meta = per_name.get(name, {}) if isinstance(per_name, dict) else {}
            if meta.get("name_type") == "alias_to_external_target":
                return (
                    "Tool found a CNAME target outside the scanned base domain. "
                    "The external target is not counted as a child domain."
                )
            if meta.get("name_type") in {"generic_hostname", "technical_vendor_hostname"}:
                record_hint = "CNAME" if meta.get("name_type") == "technical_vendor_hostname" else "A"
                return (
                    f"Tool found DNS records for {name}, but it appears to be a generic "
                    f"hostname. Optional manual check: dig {record_hint} {name}"
                )
            if meta.get("name_type") == "alias_to_known_domain":
                return (
                    f"Tool found DNS records for {name}, but it points to a child domain "
                    f"already known in the system. Optional manual check: dig CNAME {name}"
                )
        child = sorted(new_children)[0]
        return (
            f"Tool found DNS records for {child}, but it appears to be a generic hostname. "
            f"Optional manual check: dig A {child}"
        )

    if evidence_value == "validation_only":
        validated = comparison.get("known_domains_validated", "")
        if validated:
            first = str(validated).split(";")[0].strip()
            if first:
                return (
                    f"Tool confirmed a domain already listed in the system input. "
                    f"Optional manual check: dig NS {first}; dig A {first}"
                )
        return "Tool confirmed domains already listed in the system input."

    if evidence_value == "context_only" or counts["base_zone_exists_flag"]:
        return (
            f"Tool found base zone DNS evidence for {base}. "
            f"Optional manual check: dig SOA {base}; dig NS {base}"
        )

    return (
        "Tool did not find child DNS names with the selected methods. "
        "No manual check is suggested unless this domain is important."
    )


def _finding_child_metadata(
    domain_result: DomainScanResult,
    fqdn: str,
    comparison: dict[str, str | int | bool],
    *,
    record: DiscoveredRecord | None = None,
) -> dict[str, str]:
    per_name = comparison.get("_per_name", {})
    ctx = comparison.get("_dns_ctx")
    if not isinstance(ctx, DnsEvidenceContext):
        ctx = _build_dns_evidence_context(domain_result)

    normalized = normalize_domain_name(fqdn)
    if isinstance(per_name, dict) and normalized in per_name:
        meta = dict(per_name[normalized])
    elif isinstance(per_name, dict) and fqdn in per_name:
        meta = dict(per_name[fqdn])
    else:
        known = comparison.get("_known_set", set())
        known_domain = "yes" if normalized in known else "no"
        is_delegated = _has_direct_delegation(fqdn, ctx)
        known_set = known if isinstance(known, set) else set()
        name_type = _classify_name_type(
            fqdn,
            domain_result.domain,
            is_delegated=is_delegated,
            ctx=ctx,
            known=known_set,
        )
        evidence_value = _evidence_value_for_child_name(
            known_domain=known_domain == "yes",
            name_type=name_type,
            is_delegated=is_delegated,
            has_axfr_child=False,
        )
        meta = {
            "discovered_name": fqdn,
            "known_domain": known_domain,
            "name_type": name_type,
            "evidence_value": evidence_value,
        }

    if record is not None and record.record_type == RecordType.CNAME:
        target = normalize_domain_name(record.value.rstrip("."))
        known_set = comparison.get("_known_set", set())
        if not isinstance(known_set, set):
            known_set = set()
        if not _is_child_name(target, domain_result.domain):
            meta["name_type"] = "alias_to_external_target"
            meta["evidence_value"] = "limited"
            meta["known_domain"] = "no"
        elif target in known_set:
            meta["name_type"] = "alias_to_known_domain"
            meta["evidence_value"] = "limited"
            if normalized not in known_set:
                meta["known_domain"] = "no"

    if record is not None and record.classification == FindingClassification.DELEGATED_CHILD_ZONE:
        if not _has_direct_delegation(fqdn, ctx):
            meta["name_type"] = _classify_name_type(
                fqdn,
                domain_result.domain,
                is_delegated=False,
                ctx=ctx,
                known=comparison.get("_known_set") if isinstance(comparison.get("_known_set"), set) else None,
            )
            meta["evidence_value"] = _evidence_value_for_child_name(
                known_domain=meta["known_domain"] == "yes",
                name_type=meta["name_type"],
                is_delegated=False,
                has_axfr_child=False,
            )

    meta["why"] = _why_for_child_name(
        fqdn=fqdn,
        base_domain=domain_result.domain,
        known_domain=meta["known_domain"] == "yes",
        name_type=meta["name_type"],
        evidence_value=meta["evidence_value"],
        ctx=ctx,
        record=record,
    )
    return meta


def _wordlist_sources_text(result: ScanRunResult) -> str:
    if result.wordlist_plan and result.wordlist_plan.source_counts:
        return "; ".join(result.wordlist_plan.source_counts.keys())
    return ""


def _map_confidence(value: str) -> str:
    mapping = {
        "normal": "high",
        "high": "high",
        "medium": "medium",
        "low": "low",
    }
    return mapping.get(value, "unknown")


def _map_source(record: DiscoveredRecord) -> str:
    mapping = {
        "recursive_resolver": "recursive",
        "authoritative_nameserver": "authoritative",
        "axfr": "axfr",
        "wildcard_probe": "wildcard_probe",
        "generated_candidate": "generated_candidate",
        "fifth_level_parent_validation": "5th-level parent check",
    }
    return mapping.get(record.source_method, record.source_method)


def _domain_has_scan_error(domain_result: DomainScanResult) -> bool:
    if domain_result.scan_failed:
        return True
    if any(record.classification == FindingClassification.SCAN_ERROR for record in domain_result.records):
        return True
    return any(
        "unexpected error" in note.lower() or "interrupted" in note.lower()
        for note in domain_result.notes
    )


def _is_dns_activity_record(record: DiscoveredRecord) -> bool:
    if record.classification == FindingClassification.DELEGATED_CHILD_ZONE:
        return False
    if record.classification in {
        FindingClassification.BASE_ZONE_EXISTS,
        FindingClassification.ZONE_SOA_DISCOVERED,
        FindingClassification.STANDARD_RECORD,
    }:
        return True
    if record.classification == FindingClassification.BASE_DOMAIN_RECORD:
        return record.record_type is not None and record.record_type.value != "NS"
    return False


def _domain_summary_counts(domain_result: DomainScanResult) -> dict[str, int]:
    counts = {
        "total_findings": 0,
        "base_domain_records": 0,
        "authoritative_ns": 0,
        "base_zone_exists": 0,
        "delegated_child_zones": 0,
        "dns_names_with_records": 0,
        "standard_records": 0,
        "axfr_success": 0,
        "axfr_blocked": 0,
        "query_errors": 0,
        "scan_errors": 0,
        "candidates_tested": domain_result.candidates_tested,
    }
    dns_activity_names: set[str] = set()
    delegated_names: set[str] = set()
    base_has_zone = False

    for record in domain_result.records:
        counts["total_findings"] += 1
        key_map = {
            FindingClassification.BASE_DOMAIN_RECORD: "base_domain_records",
            FindingClassification.BASE_ZONE_EXISTS: "base_zone_exists",
            FindingClassification.AUTHORITATIVE_NS: "authoritative_ns",
            FindingClassification.DELEGATED_CHILD_ZONE: "delegated_child_zones",
            FindingClassification.STANDARD_RECORD: "standard_records",
            FindingClassification.AXFR_SUCCESS: "axfr_success",
            FindingClassification.AXFR_BLOCKED: "axfr_blocked",
            FindingClassification.QUERY_ERROR: "query_errors",
            FindingClassification.SCAN_ERROR: "scan_errors",
        }
        bucket = key_map.get(record.classification)
        if bucket:
            counts[bucket] += 1
        if record.classification == FindingClassification.ZONE_SOA_DISCOVERED:
            if record.fqdn == domain_result.domain:
                counts["base_zone_exists"] += 1

        if record.classification == FindingClassification.DELEGATED_CHILD_ZONE:
            delegated_names.add(record.fqdn)
        elif _is_dns_activity_record(record):
            dns_activity_names.add(record.fqdn)

        if record.classification == FindingClassification.BASE_ZONE_EXISTS and record.fqdn == domain_result.domain:
            base_has_zone = True
        if (
            record.record_type is not None
            and record.record_type.value == "SOA"
            and record.fqdn == domain_result.domain
            and record.classification
            in {
                FindingClassification.BASE_ZONE_EXISTS,
                FindingClassification.BASE_DOMAIN_RECORD,
                FindingClassification.ZONE_SOA_DISCOVERED,
            }
        ):
            base_has_zone = True

    counts["delegated_child_zones"] = len(delegated_names)
    counts["dns_names_with_records"] = len(dns_activity_names)
    counts["base_zone_exists_flag"] = 1 if base_has_zone else 0
    return counts


def _axfr_status_values(domain_result: DomainScanResult, axfr_enabled: bool) -> list[str]:
    if not axfr_enabled:
        return ["not_attempted"]

    statuses: list[str] = []
    if any(record.classification == FindingClassification.AXFR_SUCCESS for record in domain_result.records):
        statuses.append("success")

    for record in domain_result.records:
        if record.classification != FindingClassification.AXFR_BLOCKED:
            continue
        message = record.value.lower()
        if "timeout" in message:
            statuses.append("timeout")
        elif "refused" in message:
            statuses.append("refused")
        elif "blocked" in message or "transfererror" in message:
            statuses.append("blocked")
        else:
            statuses.append("failed")

    return statuses or ["not_attempted"]


def _domain_axfr_status(domain_result: DomainScanResult, axfr_enabled: bool) -> str:
    statuses = _axfr_status_values(domain_result, axfr_enabled)
    unique = list(dict.fromkeys(statuses))
    if len(unique) > 1:
        return "mixed"
    return unique[0]


def _authoritative_nameservers(domain_result: DomainScanResult) -> list[str]:
    names: list[str] = []
    for record in domain_result.records:
        if record.classification == FindingClassification.AUTHORITATIVE_NS and record.record_type:
            names.append(record.value.rstrip("."))

    if not names:
        for record in domain_result.records:
            if record.fqdn != domain_result.domain:
                continue
            if record.record_type is None or record.record_type.value != "SOA":
                continue
            if record.classification not in {
                FindingClassification.BASE_ZONE_EXISTS,
                FindingClassification.BASE_DOMAIN_RECORD,
                FindingClassification.ZONE_SOA_DISCOVERED,
            }:
                continue
            mname = record.value.split()[0].rstrip(".")
            if mname:
                names.append(f"{mname} (authoritative indicator from SOA)")

    return list(dict.fromkeys(names))


def _should_emit_no_records_row(domain_result: DomainScanResult, counts: dict[str, int]) -> bool:
    if _domain_has_scan_error(domain_result):
        return False
    if counts["base_zone_exists_flag"]:
        return False
    if counts["delegated_child_zones"] > 0 or counts["dns_names_with_records"] > 0:
        return False
    if counts["base_domain_records"] > 0 or counts["authoritative_ns"] > 0:
        return False
    if counts["axfr_success"] > 0:
        return False
    return True


def _include_record_in_export(record: DiscoveredRecord, base_domain: str) -> bool:
    if record.classification in {
        FindingClassification.QUERY_ERROR,
        FindingClassification.SCAN_ERROR,
    }:
        return record.fqdn == base_domain
    return True


def _record_type_text(record: DiscoveredRecord) -> str:
    if record.classification == FindingClassification.AXFR_SUCCESS and record.record_type is None:
        return "AXFR"
    if record.record_type is None:
        return ""
    return record.record_type.value


def _record_error(record: DiscoveredRecord) -> str:
    if record.classification in {
        FindingClassification.QUERY_ERROR,
        FindingClassification.SCAN_ERROR,
    }:
        return record.value
    return ""


def _finding_notes(record: DiscoveredRecord, base_notes: str) -> str:
    if record.source_method == "fifth_level_parent_validation":
        return f"{FIFTH_LEVEL_PARENT_NOTE} {base_notes}".strip()
    if record.classification in {
        FindingClassification.BASE_ZONE_EXISTS,
        FindingClassification.ZONE_SOA_DISCOVERED,
    }:
        return f"{SOA_FINDING_NOTE} {base_notes}".strip()
    if record.classification == FindingClassification.SCAN_ERROR:
        return f"{SCAN_ERROR_NOTE} {base_notes}".strip()
    return base_notes


def _record_value(record: DiscoveredRecord) -> str:
    if record.classification == FindingClassification.QUERY_ERROR:
        return ""
    if record.classification == FindingClassification.NO_RECORDS_DISCOVERED:
        return "No records discovered using tested methods"
    return record.value


def _base_has_discovered_records(domain_result: DomainScanResult) -> bool:
    for record in domain_result.records:
        if record.fqdn != domain_result.domain:
            continue
        if record.classification in {
            FindingClassification.BASE_DOMAIN_RECORD,
            FindingClassification.AUTHORITATIVE_NS,
            FindingClassification.BASE_ZONE_EXISTS,
            FindingClassification.ZONE_SOA_DISCOVERED,
        }:
            return True
    return False


def _determine_scan_status(domain_result: DomainScanResult, counts: dict[str, int]) -> str:
    has_error = _domain_has_scan_error(domain_result)
    has_delegated = counts["delegated_child_zones"] > 0
    has_dns = counts["dns_names_with_records"] > 0
    has_base_zone = bool(counts["base_zone_exists_flag"])
    has_base_records = counts["base_domain_records"] > 0 or counts["authoritative_ns"] > 0
    has_meaningful = (
        has_delegated
        or has_dns
        or has_base_zone
        or has_base_records
        or counts["axfr_success"] > 0
    )

    if counts["axfr_success"] > 0:
        return SCAN_STATUS_AXFR
    if has_delegated:
        return SCAN_STATUS_DELEGATED_CHILD
    if has_error and has_dns:
        return SCAN_STATUS_DNS_ACTIVITY_WITH_ERRORS
    if has_dns:
        return SCAN_STATUS_DNS_ACTIVITY
    if has_base_zone:
        return SCAN_STATUS_BASE_ZONE_EXISTS
    if has_base_records:
        return SCAN_STATUS_BASE_ONLY
    if has_error:
        if has_meaningful or domain_result.candidates_tested > 0:
            return SCAN_STATUS_INCOMPLETE
        return SCAN_STATUS_INCOMPLETE if domain_result.scan_failed else SCAN_STATUS_ERRORS_ONLY
    if counts["query_errors"] > 0 and counts["total_findings"] == counts["query_errors"]:
        return SCAN_STATUS_ERRORS_ONLY
    return SCAN_STATUS_NO_RECORDS


def _evidence_summary(domain_result: DomainScanResult, counts: dict[str, int]) -> str:
    parts: list[str] = []

    if _domain_has_scan_error(domain_result):
        parts.append("Scan incomplete due to error; rerun recommended.")

    if counts["axfr_success"] > 0:
        parts.append(f"AXFR succeeded with {counts['axfr_success']} record(s) discovered using tested methods")

    delegated = [
        record.fqdn
        for record in domain_result.records
        if record.classification == FindingClassification.DELEGATED_CHILD_ZONE
    ]
    if delegated:
        unique = list(dict.fromkeys(delegated))[:5]
        suffix = "..." if len(delegated) > 5 else ""
        sample = unique[0]
        if len(unique) == 1:
            parts.append(f"Delegated child zone discovered: {sample} has NS records.")
        else:
            parts.append(f"Delegated child zones discovered: {', '.join(unique)}{suffix}")

    dns_names = sorted(
        {
            record.fqdn
            for record in domain_result.records
            if _is_dns_activity_record(record) and not (
                record.fqdn == domain_result.domain
                and record.classification == FindingClassification.BASE_ZONE_EXISTS
            )
        }
    )
    if dns_names:
        examples = dns_names[:3]
        suffix = "..." if len(dns_names) > 3 else ""
        parts.append(
            f"DNS activity discovered: {', '.join(examples)}{suffix} has A/CNAME/MX/TXT/SOA or related records."
        )

    base_soa = [
        record
        for record in domain_result.records
        if record.fqdn == domain_result.domain
        and record.record_type is not None
        and record.record_type.value == "SOA"
    ]
    apex_a = [
        record
        for record in domain_result.records
        if record.fqdn == domain_result.domain
        and record.record_type is not None
        and record.record_type.value == "A"
    ]
    if base_soa and not apex_a:
        parts.append("SOA discovered for base zone; no apex A record found.")
    elif base_soa:
        parts.append("SOA discovered for base zone.")

    base_records = [
        record
        for record in domain_result.records
        if record.classification == FindingClassification.BASE_DOMAIN_RECORD
        and record.fqdn == domain_result.domain
        and record.record_type is not None
        and record.record_type.value != "SOA"
    ]
    if base_records:
        examples = list(
            dict.fromkeys(
                f"{item.record_type.value if item.record_type else 'record'}={item.value[:60]}"
                for item in base_records
            )
        )[:4]
        parts.append(f"Base domain records: {', '.join(examples)}")

    if domain_result.notes:
        for note in domain_result.notes:
            if "authoritative indicator from soa" in note.lower():
                parts.append(note)
            elif "No records discovered" in note and not parts:
                parts.append(note)

    if not parts:
        return "No records discovered using selected wordlists and tested methods."
    return "; ".join(dict.fromkeys(parts))


def build_findings_rows(result: ScanRunResult) -> list[dict[str, str]]:
    """Flatten scan results into findings row dictionaries."""
    return build_csv_rows(result)


def build_csv_rows(result: ScanRunResult) -> list[dict[str, str]]:
    """Flatten scan results into CSV row dictionaries."""
    rows: list[dict[str, str]] = []
    scan_timestamp = _format_timestamp(result.scan_timestamp)
    wordlist_sources = _wordlist_sources_text(result)
    axfr_enabled = result.input.options.attempt_axfr
    notes = _export_notes(result)

    for domain_result in result.domain_results:
        axfr_status = _domain_axfr_status(domain_result, axfr_enabled)
        wildcard = "true" if domain_result.wildcard_suspected else "false"
        counts = _domain_summary_counts(domain_result)
        comparison = _compare_known_vs_discovered(domain_result)

        if _should_emit_no_records_row(domain_result, counts):
            domain_why = _why_for_domain_summary(
                str(comparison.get("evidence_value", "none")),
                comparison,
                scan_failed=_domain_has_scan_error(domain_result),
            )
            rows.append(
                {
                    "scan_timestamp": scan_timestamp,
                    "base_domain": domain_result.domain,
                    "discovered_name": domain_result.domain,
                    "known_domain": "yes",
                    "name_type": "base_domain",
                    "evidence_value": str(comparison.get("evidence_value", "none")),
                    "why": domain_why,
                    "tested_name": domain_result.domain,
                    "record_type": "",
                    "finding_type": FindingClassification.NO_RECORDS_DISCOVERED.value,
                    "confidence": "unknown",
                    "source": "recursive",
                    "nameserver": "",
                    "value": "No records discovered using tested methods",
                    "ttl": "",
                    "wildcard_suspected": wildcard,
                    "axfr_status": axfr_status,
                    "error": "",
                    "wordlist_sources": wordlist_sources,
                    "notes": notes,
                }
            )

        for record in domain_result.records:
            if not _include_record_in_export(record, domain_result.domain):
                continue
            child_meta = _finding_child_metadata(
                domain_result,
                record.fqdn,
                comparison,
                record=record,
            )

            rows.append(
                {
                    "scan_timestamp": scan_timestamp,
                    "base_domain": domain_result.domain,
                    **child_meta,
                    "tested_name": record.fqdn,
                    "record_type": _record_type_text(record),
                    "finding_type": record.classification.value,
                    "confidence": _map_confidence(record.confidence),
                    "source": _map_source(record),
                    "nameserver": record.nameserver or "",
                    "value": _record_value(record),
                    "ttl": "" if record.ttl is None else str(record.ttl),
                    "wildcard_suspected": wildcard,
                    "axfr_status": axfr_status,
                    "error": _record_error(record),
                    "wordlist_sources": wordlist_sources,
                    "notes": _finding_notes(record, notes),
                }
            )

    return rows


def build_summary_rows(result: ScanRunResult) -> list[dict[str, str]]:
    """Build one summary row per base domain."""
    rows: list[dict[str, str]] = []
    scan_timestamp = _format_timestamp(result.scan_timestamp)
    wordlist_sources = _wordlist_sources_text(result)
    axfr_enabled = result.input.options.attempt_axfr

    for domain_result in result.domain_results:
        counts = _domain_summary_counts(domain_result)
        metadata = _input_metadata_values(domain_result)
        comparison = _compare_known_vs_discovered(domain_result)
        scan_status = _determine_scan_status(domain_result, counts)
        comparison_export = {
            key: str(value)
            for key, value in comparison.items()
            if not str(key).startswith("_")
        }
        evidence_value = str(comparison.get("evidence_value", "none"))
        scan_failed = _domain_has_scan_error(domain_result)
        rows.append(
            {
                "scan_timestamp": scan_timestamp,
                "base_domain": domain_result.domain,
                **metadata,
                **comparison_export,
                "evidence_value": evidence_value,
                "why": _why_for_domain_summary(
                    evidence_value,
                    comparison,
                    scan_failed=scan_failed,
                ),
                "recommended_review_action": _recommended_review_action(evidence_value),
                "manual_verification_hint": _manual_verification_hint(
                    domain_result,
                    counts,
                    scan_status,
                    comparison,
                    evidence_value,
                ),
                "scan_status": scan_status,
                "authoritative_nameservers": "; ".join(_authoritative_nameservers(domain_result)),
                "axfr_status": _domain_axfr_status(domain_result, axfr_enabled),
                "wildcard_suspected": "true" if domain_result.wildcard_suspected else "false",
                "base_zone_exists": "true" if counts["base_zone_exists_flag"] else "false",
                "candidate_names_tested": str(counts["candidates_tested"]),
                "wordlist_sources": wordlist_sources,
                "analysis_note": _analysis_note(domain_result, counts, scan_status, comparison),
                "limitation_note": _limitation_note(result),
            }
        )

    return rows


def build_evidence_review_rows(result: ScanRunResult) -> list[dict[str, str]]:
    """Build coworker-ready evidence review rows sorted by support level."""
    summary_rows = build_summary_rows(result)
    review_rows = [
        {column: row.get(column, "") for column in EVIDENCE_REVIEW_COLUMNS}
        for row in summary_rows
    ]
    review_rows.sort(
        key=lambda row: EVIDENCE_VALUE_ORDER.get(row.get("evidence_value", "none"), 99)
    )
    return review_rows


def build_settings_rows(result: ScanRunResult) -> list[tuple[str, str]]:
    """Build key/value rows for the Scan Settings sheet."""
    plan = result.wordlist_plan
    options = result.input.options
    label_counts = plan.source_counts if plan else {}

    load_info = result.input_load_info
    rows: list[tuple[str, str]] = [
        ("app_name", APP_NAME),
        ("scan_timestamp", _format_timestamp(result.scan_timestamp)),
        ("export_timestamp", _format_timestamp(result.finished_at or result.scan_timestamp)),
        ("output_folder", str(result.input.output_dir.resolve()) if result.input.output_dir else ""),
        ("packaged_mode", str(is_frozen()).lower()),
        ("scan_profile", options.scan_profile.value),
        ("evidence_model_version", EVIDENCE_MODEL_VERSION),
        ("child_domain_discovery_goal", CHILD_DOMAIN_DISCOVERY_GOAL),
        ("known_domain_definition", KNOWN_DOMAIN_DEFINITION),
        ("evidence_value_definition", EVIDENCE_VALUE_DEFINITION),
        ("domain_definition_note", DOMAIN_DEFINITION_NOTE),
        ("input_file_type", load_info.input_file_type if load_info else ""),
        ("preferred_input_format_detected", str(load_info.preferred_input_format_detected).lower() if load_info else "false"),
        ("recommended_input_columns", RECOMMENDED_INPUT_COLUMNS_CSV),
        ("recommended_input_format_note", PREFERRED_INPUT_FORMAT_NOTE),
        ("metadata_columns_detected", ", ".join(load_info.metadata_columns_detected) if load_info else ""),
        ("domains_loaded", str(load_info.domains_loaded if load_info else len(result.domain_inputs))),
        ("duplicate_domains_removed", str(load_info.duplicate_domains_removed if load_info else 0)),
        (
            "input_metadata_preserved",
            str(load_info.input_metadata_preserved).lower() if load_info else "false",
        ),
        ("domains_scanned", str(len(result.domain_results))),
        ("domains_planned", str(result.domains_total or len(result.domains_planned))),
        ("scan_completed", str(result.scan_status == ScanStatus.COMPLETED).lower()),
        ("scan_cancelled", str(result.cancelled).lower()),
        ("partial_results", str(result.partial).lower()),
        ("elapsed_seconds", "" if result.elapsed_seconds is None else f"{result.elapsed_seconds:.1f}"),
        ("selected_wordlist_sources", _wordlist_sources_text(result)),
        ("label_count_by_source", json.dumps(label_counts)),
        ("total_unique_labels", str(plan.total_unique_labels if plan else 0)),
        ("estimated_candidates_per_domain", str(plan.estimated_candidates_per_domain if plan else 0)),
        ("axfr_enabled", str(options.attempt_axfr).lower()),
        ("authoritative_query_enabled", str(options.query_authoritative_ns).lower()),
        ("custom_wordlist_used", str(options.include_custom_wordlist).lower()),
        ("dns_timeout", str(DNS_TIMEOUT)),
        ("dns_lifetime", str(DNS_LIFETIME)),
        (
            "fifth_level_parent_validation",
            str(
                bool(
                    plan
                    and (
                        plan.fifth_level_enabled
                        or plan.known_fifth_level_candidates > 0
                    )
                )
            ).lower(),
        ),
        (
            "fifth_level_parent_validation_note",
            (
                "Parent 4th-level names are checked once when deeper candidate names are tested."
                if plan and (plan.fifth_level_enabled or plan.known_fifth_level_candidates > 0)
                else ""
            ),
        ),
        ("discovery_limitation", DISCOVERY_LIMITATION),
        ("context_limitation", CONTEXT_LIMITATION),
        ("partial_scan_note", PARTIAL_SCAN_NOTE if result.partial or result.cancelled else ""),
        ("operator_note", OPERATOR_NOTE),
        ("review_path_note", REVIEW_PATH_NOTE),
    ]
    return rows


def _warning_context(domain_result: DomainScanResult | None) -> dict[str, str]:
    if domain_result is None:
        return {"delegated_manager": "", "zone": ""}
    meta = _input_metadata_values(domain_result)
    return {
        "delegated_manager": meta["delegated_manager"],
        "zone": meta["zone"],
    }


def build_errors_warning_rows(result: ScanRunResult) -> list[dict[str, str]]:
    """Build domain-level warnings and errors for export."""
    rows: list[dict[str, str]] = []
    scan_timestamp = _format_timestamp(result.scan_timestamp)
    axfr_enabled = result.input.options.attempt_axfr
    plan = result.wordlist_plan

    if result.partial or result.cancelled:
        rows.append(
            {
                "scan_timestamp": scan_timestamp,
                "base_domain": "",
                "delegated_manager": "",
                "zone": "",
                "warning_type": "scan_cancelled",
                "tested_name": "",
                "record_type": "",
                "nameserver": "",
                "message": PARTIAL_SCAN_NOTE,
                "notes": _export_notes(result),
            }
        )

    if plan and plan.estimated_candidates_per_domain > CANDIDATE_STRONG_WARN_THRESHOLD:
        rows.append(
            {
                "scan_timestamp": scan_timestamp,
                "base_domain": "",
                "delegated_manager": "",
                "zone": "",
                "warning_type": "large_candidate_count",
                "tested_name": "",
                "record_type": "",
                "nameserver": "",
                "message": (
                    f"WARNING: estimated {plan.estimated_candidates_per_domain} candidates per domain "
                    "is very large — scan time and noise may increase significantly."
                ),
                "notes": DISCOVERY_LIMITATION,
            }
        )
    elif plan and plan.estimated_candidates_per_domain > CANDIDATE_WARN_THRESHOLD:
        rows.append(
            {
                "scan_timestamp": scan_timestamp,
                "base_domain": "",
                "delegated_manager": "",
                "zone": "",
                "warning_type": "large_candidate_count",
                "tested_name": "",
                "record_type": "",
                "nameserver": "",
                "message": (
                    f"Warning: estimated {plan.estimated_candidates_per_domain} candidates per domain "
                    "may increase scan time and noise."
                ),
                "notes": DISCOVERY_LIMITATION,
            }
        )

    for domain_result in result.domain_results:
        context = _warning_context(domain_result)
        if domain_result.wildcard_suspected:
            rows.append(
                {
                    "scan_timestamp": scan_timestamp,
                    "base_domain": domain_result.domain,
                    **context,
                    "warning_type": "wildcard_suspected",
                    "tested_name": domain_result.domain,
                    "record_type": "",
                    "nameserver": "",
                    "message": "Wildcard suspected using tested methods; some candidate results marked lower confidence.",
                    "notes": DISCOVERY_LIMITATION,
                }
            )

        for record in domain_result.records:
            if record.classification == FindingClassification.AXFR_BLOCKED:
                message = record.value
                warning_type = "axfr_blocked"
                lower = message.lower()
                if "timeout" in lower:
                    warning_type = "axfr_timeout"
                elif "refused" in lower:
                    warning_type = "axfr_refused"
                elif "failed" in lower:
                    warning_type = "axfr_failed"
                rows.append(
                    {
                        "scan_timestamp": scan_timestamp,
                        "base_domain": domain_result.domain,
                        **context,
                        "warning_type": warning_type,
                        "tested_name": domain_result.domain,
                        "record_type": "AXFR",
                        "nameserver": record.nameserver or "",
                        "message": message,
                        "notes": DISCOVERY_LIMITATION,
                    }
                )

            if record.classification == FindingClassification.QUERY_ERROR and record.fqdn == domain_result.domain:
                rows.append(
                    {
                        "scan_timestamp": scan_timestamp,
                        "base_domain": domain_result.domain,
                        **context,
                        "warning_type": "query_error",
                        "tested_name": record.fqdn,
                        "record_type": _record_type_text(record),
                        "nameserver": record.nameserver or "",
                        "message": record.value,
                        "notes": DISCOVERY_LIMITATION,
                    }
                )

            if record.classification == FindingClassification.SCAN_ERROR and record.fqdn == domain_result.domain:
                rows.append(
                    {
                        "scan_timestamp": scan_timestamp,
                        "base_domain": domain_result.domain,
                        **context,
                        "warning_type": "scan_error",
                        "tested_name": record.fqdn,
                        "record_type": "",
                        "nameserver": "",
                        "message": record.value,
                        "notes": SCAN_ERROR_NOTE,
                    }
                )

        for note in domain_result.notes:
            if "unexpected error" in note.lower() or "interrupted" in note.lower():
                if any(
                    row.get("base_domain") == domain_result.domain
                    and row.get("warning_type") == "scan_error"
                    for row in rows
                ):
                    continue
                rows.append(
                    {
                        "scan_timestamp": scan_timestamp,
                        "base_domain": domain_result.domain,
                        **context,
                        "warning_type": "scan_error",
                        "tested_name": domain_result.domain,
                        "record_type": "",
                        "nameserver": "",
                        "message": note,
                        "notes": SCAN_ERROR_NOTE,
                    }
                )

    return rows


def _finding_to_dict(
    record: DiscoveredRecord,
    base_domain: str,
    include_errors: bool,
) -> dict | None:
    if record.classification in {
        FindingClassification.QUERY_ERROR,
        FindingClassification.SCAN_ERROR,
    }:
        if not include_errors or record.fqdn != base_domain:
            return None
        return {
            "tested_name": record.fqdn,
            "record_type": None,
            "finding_type": record.classification.value,
            "confidence": _map_confidence(record.confidence),
            "source": _map_source(record),
            "nameserver": record.nameserver,
            "value": None,
            "ttl": record.ttl,
            "error": record.value,
        }

    return {
        "tested_name": record.fqdn,
        "record_type": _record_type_text(record) or None,
        "finding_type": record.classification.value,
        "confidence": _map_confidence(record.confidence),
        "source": _map_source(record),
        "nameserver": record.nameserver,
        "value": _record_value(record) or record.value,
        "ttl": record.ttl,
        "error": None,
    }


def build_json_document(result: ScanRunResult) -> dict:
    """Build the JSON export document."""
    plan = result.wordlist_plan
    metadata = {
        "scan_timestamp": _format_timestamp(result.scan_timestamp),
        "app_name": APP_NAME,
        "app_version": APP_VERSION,
        "selected_wordlist_sources": plan.source_counts if plan else {},
        "total_unique_labels": plan.total_unique_labels if plan else 0,
        "estimated_candidates_per_domain": plan.estimated_candidates_per_domain if plan else 0,
        "axfr_enabled": result.input.options.attempt_axfr,
        "authoritative_query_enabled": result.input.options.query_authoritative_ns,
        "scan_completed": result.scan_status == ScanStatus.COMPLETED,
        "scan_cancelled": result.cancelled,
        "partial_results": result.partial,
        "elapsed_seconds": result.elapsed_seconds,
        "discovery_limitation": DISCOVERY_LIMITATION,
        "context_limitation": CONTEXT_LIMITATION,
        "partial_scan_note": PARTIAL_SCAN_NOTE if result.partial or result.cancelled else "",
    }
    if result.input_load_info:
        metadata["input_file_type"] = result.input_load_info.input_file_type
        metadata["metadata_columns_detected"] = result.input_load_info.metadata_columns_detected
        metadata["domains_loaded"] = result.input_load_info.domains_loaded
        metadata["duplicate_domains_removed"] = result.input_load_info.duplicate_domains_removed
        metadata["input_metadata_preserved"] = result.input_load_info.input_metadata_preserved

    domains: list[dict] = []
    axfr_enabled = result.input.options.attempt_axfr

    for domain_result in result.domain_results:
        findings: list[dict] = []
        errors: list[str] = []
        counts = _domain_summary_counts(domain_result)

        if _should_emit_no_records_row(domain_result, counts):
            findings.append(
                {
                    "tested_name": domain_result.domain,
                    "record_type": None,
                    "finding_type": FindingClassification.NO_RECORDS_DISCOVERED.value,
                    "confidence": "unknown",
                    "source": "recursive",
                    "nameserver": None,
                    "value": "No records discovered using tested methods",
                    "ttl": None,
                    "error": None,
                }
            )

        for record in domain_result.records:
            if record.classification in {
                FindingClassification.QUERY_ERROR,
                FindingClassification.SCAN_ERROR,
            }:
                if record.fqdn == domain_result.domain:
                    errors.append(record.value)
                continue
            item = _finding_to_dict(record, domain_result.domain, include_errors=False)
            if item:
                findings.append(item)

        comparison = _compare_known_vs_discovered(domain_result)
        scan_status = _determine_scan_status(domain_result, counts)
        comparison_export = {
            key: value
            for key, value in comparison.items()
            if not key.startswith("_")
        }
        evidence_value = str(comparison.get("evidence_value", "none"))

        domains.append(
            {
                "base_domain": domain_result.domain,
                "input_domain": domain_result.input_record.original_domain if domain_result.input_record else None,
                "delegated_manager": _input_metadata_values(domain_result)["delegated_manager"] or None,
                "zone": _input_metadata_values(domain_result)["zone"] or None,
                "child_domain_inventory": {
                    key: value
                    for key, value in comparison.items()
                    if not str(key).startswith("_")
                },
                "evidence_value": evidence_value,
                "recommended_review_action": _recommended_review_action(evidence_value),
                "manual_verification_hint": _manual_verification_hint(
                    domain_result,
                    counts,
                    scan_status,
                    comparison,
                    evidence_value,
                ),
                "analysis_note": _analysis_note(domain_result, counts, scan_status, comparison),
                "wildcard_suspected": domain_result.wildcard_suspected,
                "scan_failed": domain_result.scan_failed,
                "authoritative_nameservers": _authoritative_nameservers(domain_result),
                "axfr_status": _domain_axfr_status(domain_result, axfr_enabled),
                "scan_status": _determine_scan_status(domain_result, counts),
                "findings": findings,
                "summary_counts": counts,
                "errors": errors,
                "notes": domain_result.notes + [DISCOVERY_LIMITATION],
            }
        )

    return {
        "scan_metadata": metadata,
        "domains": domains,
    }


def build_how_to_read_rows() -> list[tuple[str, str]]:
    """Build short coworker guidance for the How to Read workbook sheet."""
    return [
        (
            "What this report is for",
            "This report checks whether live DNS contains child names under known 3rd-level "
            ".US domains that were not listed in the system input.",
        ),
        (
            "Known domain",
            "Known domain = Yes means the discovered name was already listed in the input/system "
            "data. Known domain = No means the name was found in live DNS but was not listed in "
            "the input/system data.",
        ),
        (
            "New child domains found",
            "Lists DNS names beneath the scanned base domain that were found in live DNS testing "
            "but were not already listed in the system input. Some may be service hostnames "
            "(for example www or mail) rather than separately registered domains.",
        ),
        (
            "Evidence value — Strong",
            "Strong evidence usually means a new delegated child domain was found with NS or SOA "
            "records.",
        ),
        (
            "Evidence value — Moderate",
            "Moderate evidence means a new meaningful child name was found, such as an "
            "organizational or service name.",
        ),
        (
            "Evidence value — Limited",
            "Limited evidence usually means a generic or technical hostname such as www, mail, or "
            "autodiscover, or a DNS alias pointing to a domain already known in the system.",
        ),
        (
            "Evidence value — Validation only",
            "Known domains validated are not new discoveries. They confirm names already listed "
            "in the system input.",
        ),
        (
            "Why column",
            "The Why column explains, in plain language, why each row received its evidence "
            "value or classification.",
        ),
        (
            "What this report cannot prove",
            "No records discovered does not prove no child domains exist. It only means none were "
            "found using the selected methods.",
        ),
        (
            "Suggested review path",
            "Open Evidence Review first. Focus on Strong and Moderate rows for coworker review. "
            "Read Why for context. Use Verification guidance for optional independent dig checks.",
        ),
    ]


def _write_table_sheet(
    worksheet: Worksheet,
    columns: list[SheetColumn] | list[str],
    rows: list[dict[str, str]] | list[tuple[str, str]],
    *,
    wrap_column_keys: set[str] | None = None,
    highlight_column_key: str | None = None,
    highlight_fill_map: dict[str, PatternFill] | None = None,
    status_column_key: str | None = None,
    status_fill_map: dict[str, PatternFill] | None = None,
    tuple_column_labels: tuple[str, str, str, str] | None = None,
) -> None:
    """Write a worksheet table. Tuple rows use (key, label, key, label) column metadata."""
    wrap_column_keys = wrap_column_keys or set()
    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    if rows and isinstance(rows[0], tuple):
        left_key, left_label, right_key, right_label = tuple_column_labels or (
            "setting",
            "Setting",
            "value",
            "Value",
        )
        worksheet.append([left_label, right_label])
        for row in rows:
            worksheet.append(list(row))
        sheet_columns = [
            SheetColumn(left_key, left_label),
            SheetColumn(right_key, right_label),
        ]
    elif columns and isinstance(columns[0], SheetColumn):
        sheet_columns = columns
        headers = [column.label for column in sheet_columns]
        worksheet.append(headers)
        for row in rows:
            worksheet.append([row.get(column.key, "") for column in sheet_columns])
    else:
        sheet_columns = [SheetColumn(key, key) for key in columns]
        headers = [column.label for column in sheet_columns]
        worksheet.append(headers)
        for row in rows:
            worksheet.append([row.get(column.key, "") for column in sheet_columns])

    for cell in worksheet[1]:
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    if worksheet.max_row > 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for index, column in enumerate(sheet_columns, start=1):
        column_letter = get_column_letter(index)
        max_length = len(column.label)
        for row_cells in worksheet.iter_rows(min_row=2, min_col=index, max_col=index):
            for cell in row_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 80))
                if column.key in wrap_column_keys:
                    cell.alignment = wrap_alignment
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    if highlight_column_key and highlight_fill_map:
        highlight_index = next(
            (idx + 1 for idx, col in enumerate(sheet_columns) if col.key == highlight_column_key),
            None,
        )
        if highlight_index:
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_index, column=highlight_index)
                fill = highlight_fill_map.get(str(cell.value))
                if fill:
                    for col in range(1, len(sheet_columns) + 1):
                        worksheet.cell(row=row_index, column=col).fill = fill

    if status_column_key and status_fill_map:
        status_index = next(
            (idx + 1 for idx, col in enumerate(sheet_columns) if col.key == status_column_key),
            None,
        )
        if status_index:
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_index, column=status_index)
                fill = status_fill_map.get(str(cell.value))
                if fill:
                    for col in range(1, len(sheet_columns) + 1):
                        worksheet.cell(row=row_index, column=col).fill = fill


def export_xlsx_report(result: ScanRunResult, output_dir: Path) -> Path:
    """Write the operator-facing XLSX workbook."""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{_workbook_report_stem(result.scan_timestamp)}.xlsx"

    workbook = Workbook()
    review_sheet = workbook.active
    review_sheet.title = "Evidence Review"

    review_rows = [_coworker_display_row(row) for row in build_evidence_review_rows(result)]
    _write_table_sheet(
        review_sheet,
        EVIDENCE_REVIEW_SHEET_COLUMNS,
        review_rows,
        wrap_column_keys={
            "new_child_domains_found",
            "new_delegated_domains_found",
            "new_generic_hostnames_found",
            "new_technical_hostnames_found",
            "new_organizational_domains_found",
            "known_domains_validated",
            "why",
            "manual_verification_hint",
            "limitation_note",
        },
        highlight_column_key="evidence_value",
        highlight_fill_map=EVIDENCE_VALUE_FILLS,
    )

    how_to_read_sheet = workbook.create_sheet("How to Read", 1)
    _write_table_sheet(
        how_to_read_sheet,
        [],
        build_how_to_read_rows(),
        wrap_column_keys={"value"},
        tuple_column_labels=("topic", "Topic", "value", "Explanation"),
    )

    summary_sheet = workbook.create_sheet("Summary")
    summary_rows = [_coworker_display_row(row) for row in build_summary_rows(result)]
    _write_table_sheet(
        summary_sheet,
        SUMMARY_SHEET_COLUMNS,
        summary_rows,
        wrap_column_keys={
            "why",
            "analysis_note",
            "limitation_note",
            "manual_verification_hint",
            "authoritative_nameservers",
            "wordlist_sources",
            "known_domains_from_system",
            "known_domains_validated",
            "new_child_domains_found",
            "new_delegated_domains_found",
            "new_generic_hostnames_found",
            "new_technical_hostnames_found",
            "new_organizational_domains_found",
        },
        highlight_column_key="evidence_value",
        highlight_fill_map=EVIDENCE_VALUE_FILLS,
        status_column_key="scan_status",
        status_fill_map=SUMMARY_STATUS_FILLS,
    )

    findings_sheet = workbook.create_sheet("Findings")
    findings_rows = [_coworker_display_row(row) for row in build_findings_rows(result)]
    _write_table_sheet(
        findings_sheet,
        FINDINGS_SHEET_COLUMNS,
        findings_rows,
        wrap_column_keys={"value", "why", "notes", "error", "nameserver"},
    )

    settings_sheet = workbook.create_sheet("Scan Settings")
    _write_table_sheet(
        settings_sheet,
        [],
        build_settings_rows(result),
        wrap_column_keys={"value"},
    )

    warnings_sheet = workbook.create_sheet("Errors Warnings")
    warning_rows = build_errors_warning_rows(result)
    if not warning_rows:
        warning_rows = [
            {
                "scan_timestamp": _format_timestamp(result.scan_timestamp),
                "base_domain": "",
                "warning_type": "none",
                "tested_name": "",
                "record_type": "",
                "nameserver": "",
                "message": "No domain-level warnings or errors recorded for this scan.",
                "notes": DISCOVERY_LIMITATION,
            }
        ]
    _write_table_sheet(
        warnings_sheet,
        ERRORS_WARNINGS_SHEET_COLUMNS,
        warning_rows,
        wrap_column_keys={"message", "notes"},
    )

    workbook.save(path)
    return path


def export_csv(result: ScanRunResult, output_dir: Path) -> tuple[Path, int]:
    """Write findings CSV report and return path plus row count."""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{_findings_report_stem(result.scan_timestamp)}.csv"
    rows = build_csv_rows(result)

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    return path, len(rows)


def export_summary_csv(result: ScanRunResult, output_dir: Path) -> Path:
    """Write summary CSV report."""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{_findings_report_stem(result.scan_timestamp)}_summary.csv"
    rows = build_summary_rows(result)

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=SUMMARY_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    return path


def export_json(result: ScanRunResult, output_dir: Path) -> Path:
    """Write JSON report and return path."""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{_findings_report_stem(result.scan_timestamp)}.json"
    document = build_json_document(result)

    with path.open("w", encoding="utf-8") as handle:
        json.dump(document, handle, indent=2)
        handle.write("\n")

    return path


def export_results(
    result: ScanRunResult,
    output_dir: Path,
    export_format: ExportFormat,
) -> ExportOutcome:
    """Export scan results to XLSX, CSV, JSON, or all formats."""
    findings_rows = build_findings_rows(result)
    outcome = ExportOutcome(
        domain_count=len(result.domain_results),
        row_count=len(findings_rows),
    )

    if export_format in {"xlsx", "all"}:
        outcome.xlsx_path = export_xlsx_report(result, output_dir)

    if export_format in {"csv", "all"}:
        outcome.csv_path, _ = export_csv(result, output_dir)

    if export_format in {"json", "all"}:
        outcome.json_path = export_json(result, output_dir)

    if export_format == "all":
        outcome.summary_csv_path = export_summary_csv(result, output_dir)

    return outcome
